import sqlite3
from datetime import datetime, time
from io import BytesIO

import pytest
from openpyxl import load_workbook

from src.outsource_attendance import AttendanceService, IST, classify_shift


def test_classify_shift_boundaries():
    assert classify_shift(time(7, 0))[0] == "M"
    assert classify_shift(time(8, 59))[0] == "M"
    assert classify_shift(time(9, 0))[0] == "G"
    assert classify_shift(time(11, 59))[0] == "G"
    assert classify_shift(time(12, 59))[0] == "G"
    assert classify_shift(time(13, 0))[0] == "E"
    assert classify_shift(time(16, 59))[0] == "E"
    assert classify_shift(time(19, 0))[0] == "N"
    assert classify_shift(time(21, 59))[0] == "N"
    assert classify_shift(time(22, 0))[0] == "O"


def test_observer_decision_and_admin_override(tmp_path):
    service = AttendanceService(tmp_path / "attendance.sqlite")
    outsource_id = service.add_user("Asha Contractor", "outsource", "9876543210")
    service.add_user("Observer One", "observer", "9876543211")

    entry_id = service.submit_login(
        outsource_id,
        "pc-01",
        login_at=datetime(2026, 5, 1, 8, 30, tzinfo=IST),
    )
    service.decide_entry(entry_id, "accepted", "observer", "Observer One")
    service.decide_entry(entry_id, "rejected", "admin", "Admin", "Wrong PC")

    entries = service.list_entries(month="2026-05")

    assert entries.iloc[0]["observer_status"] == "accepted"
    assert entries.iloc[0]["admin_status"] == "rejected"
    assert entries.iloc[0]["effective_status"] == "rejected"
    assert entries.iloc[0]["decision_source"] == "Admin Override"
    assert entries.iloc[0]["final_remarks"] == "Wrong PC"


def test_observer_cannot_change_completed_observer_decision(tmp_path):
    service = AttendanceService(tmp_path / "attendance.sqlite")
    outsource_id = service.add_user("Asha Contractor", "outsource", "9876543210")
    entry_id = service.submit_login(
        outsource_id,
        "pc-01",
        login_at=datetime(2026, 5, 1, 8, 30, tzinfo=IST),
    )

    service.decide_entry(entry_id, "accepted", "observer", "Observer One")

    with pytest.raises(ValueError, match="pending entries"):
        service.decide_entry(entry_id, "rejected", "observer", "Observer Two")

    entries = service.list_entries(month="2026-05")
    assert entries.iloc[0]["observer_status"] == "accepted"
    assert entries.iloc[0]["effective_status"] == "accepted"


def test_monthly_attendance_counts_only_accepted_days(tmp_path):
    service = AttendanceService(tmp_path / "attendance.sqlite")
    outsource_id = service.add_user("Ravi Outsource", "outsource", "9876543212")

    first = service.submit_login(
        outsource_id,
        "pc-a",
        login_at=datetime(2026, 5, 2, 9, 15, tzinfo=IST),
    )
    second = service.submit_login(
        outsource_id,
        "pc-b",
        login_at=datetime(2026, 5, 2, 19, 30, tzinfo=IST),
    )
    third = service.submit_login(
        outsource_id,
        "pc-c",
        login_at=datetime(2026, 5, 3, 13, 30, tzinfo=IST),
    )

    service.decide_entry(first, "accepted", "admin", "Admin")
    service.decide_entry(second, "accepted", "admin", "Admin")
    service.decide_entry(third, "rejected", "admin", "Admin")

    matrix = service.build_monthly_attendance_df("2026-05")
    row = matrix[matrix["Name"] == "Ravi Outsource"].iloc[0]

    assert row["02 Sat"] == "G/N"
    assert row["03 Sun"] == ""
    assert row["Total Present Days"] == 1


def test_pending_filter_only_includes_entries_with_no_decision(tmp_path):
    service = AttendanceService(tmp_path / "attendance.sqlite")
    outsource_id = service.add_user("Pending Filter User", "outsource", "9876543218")
    decided_entry = service.submit_login(
        outsource_id,
        "pc-a",
        login_at=datetime(2026, 5, 5, 9, 30, tzinfo=IST),
    )
    pending_entry = service.submit_login(
        outsource_id,
        "pc-b",
        login_at=datetime(2026, 5, 6, 9, 30, tzinfo=IST),
    )

    service.decide_entry(decided_entry, "accepted", "observer", "Observer One")

    pending = service.list_entries(month="2026-05", status_filter="pending")

    assert pending["id"].tolist() == [pending_entry]


def test_export_workbook_contains_attendance_sheets(tmp_path):
    service = AttendanceService(tmp_path / "attendance.sqlite")
    outsource_id = service.add_user("Neha Vendor", "outsource", "9876543213")
    entry_id = service.submit_login(
        outsource_id,
        "pc-x",
        login_at=datetime(2026, 5, 4, 7, 30, tzinfo=IST),
    )
    service.decide_entry(entry_id, "accepted", "admin", "Admin")

    workbook_bytes = service.export_attendance_workbook("2026-05")
    workbook = load_workbook(BytesIO(workbook_bytes))

    assert workbook.sheetnames == [
        "Login Register",
        "Monthly Attendance",
        "Daily Summary",
        "User Master",
        "Audit Log",
    ]
    assert workbook["Monthly Attendance"]["A1"].value == "Monthly Attendance - Outsource Attendance"
    assert workbook["Monthly Attendance"]["A4"].value == "Neha Vendor"


def test_user_authentication_uses_password_and_keeps_hash_hidden(tmp_path):
    service = AttendanceService(tmp_path / "attendance.sqlite")
    user_id = service.add_user(
        "Secure Observer",
        "observer",
        "9876543214",
        password="secret-pass",
        designation="Reviewer",
    )

    assert service.authenticate_user(user_id, "wrong", "observer") is None

    auth = service.authenticate_user(user_id, "secret-pass", "observer")
    assert auth["name"] == "Secure Observer"
    assert auth["mobile"] == "9876543214"

    users = service.list_users(role="observer")
    assert "password_hash" not in users.columns
    assert users.iloc[0]["has_password"]


def test_update_observer_profile_can_reset_password(tmp_path):
    service = AttendanceService(tmp_path / "attendance.sqlite")
    user_id = service.add_user("Observer User", "observer", "9876543215", password="old-pass")

    service.update_user_profile(
        user_id=user_id,
        name="Observer User",
        mobile="9876543215",
        role="observer",
        password="new-pass",
    )

    assert service.authenticate_user(user_id, "old-pass", "observer") is None
    assert service.authenticate_user(user_id, "new-pass", "observer")["id"] == user_id


def test_outsource_user_uses_mobile_verification_without_password(tmp_path):
    service = AttendanceService(tmp_path / "attendance.sqlite")
    user_id = service.add_user("Vendor User", "outsource", "9876543217", password="ignored")

    users = service.list_users(role="outsource")

    assert not users.iloc[0]["has_password"]
    assert service.authenticate_user(user_id, "ignored", "outsource") is None
    assert service.authenticate_outsource_user(user_id, "9876543217")["name"] == "Vendor User"


def test_existing_old_user_table_is_migrated(tmp_path):
    db_path = tmp_path / "attendance.sqlite"
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            CREATE TABLE users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                normalized_name TEXT NOT NULL,
                role TEXT NOT NULL,
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                created_by TEXT,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            INSERT INTO users (
                name, normalized_name, role, active, created_at, created_by, updated_at
            )
            VALUES ('Old User', 'old user', 'outsource', 1, '2026-05-01', 'Admin', '2026-05-01')
            """
        )

    service = AttendanceService(db_path)
    service.update_user_profile(
        user_id=1,
        name="Old User",
        mobile="9876543216",
        role="outsource",
    )

    assert service.authenticate_outsource_user(1, "9876543216")["name"] == "Old User"
