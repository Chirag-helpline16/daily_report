import io
import textwrap

from openpyxl import load_workbook

from src.mo_finder import process_mo_report_upload


def test_process_mo_report_upload_builds_same_report_workbook():
    csv_text = textwrap.dedent(
        """\
        Acknowledgement No.,Crime Aditional Information,Category,Sub Category,Status,Amount
        311001,WITHOUT OTP fraud,Financial,OTP,Open,1000
        311002,without otp fruad,Financial,OTP,Open,2000
        311003,job fraud,Employment,Job,Closed,500
        411001,shopping fraud,Ecom,Shopping,Open,900
        """
    )

    result = process_mo_report_upload(
        raw_bytes=csv_text.encode("utf-8"),
        filename="AdditionalInformationReport.csv",
        acknowledgement_prefix="311",
    )

    assert result.filename == "MO_Unique_Report_311.xlsx"
    assert result.total_rows == 4
    assert result.rows_counted == 3
    assert result.rows_ignored == 1
    assert result.unique_mo_categories == 2
    assert result.top_10[:2] == [
        {"mo": "Without OTP Fraud", "count": 2},
        {"mo": "Job Fraud", "count": 1},
    ]

    workbook = load_workbook(io.BytesIO(result.workbook), data_only=True)
    try:
        assert workbook.sheetnames == ["MO Summary", "Detail Mapping", "Rules"]
        summary = workbook["MO Summary"]
        assert summary["B6"].value == 4
        assert summary["B7"].value == 3
        assert summary["B8"].value == 1
        assert summary["B9"].value == 2

        detail = workbook["Detail Mapping"]
        assert detail.max_row == 4
        assert [detail.cell(row, 2).value for row in range(2, 5)] == [
            "Without OTP Fraud",
            "Without OTP Fraud",
            "Job Fraud",
        ]
    finally:
        workbook.close()
