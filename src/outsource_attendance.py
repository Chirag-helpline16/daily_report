"""Outsource attendance pages backed by SQLite locally or MongoDB online."""

from __future__ import annotations

import calendar
import hashlib
import hmac
import os
import sqlite3
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from pymongo import ASCENDING, DESCENDING, MongoClient, ReturnDocument
    from pymongo.errors import DuplicateKeyError, PyMongoError
except ImportError:  # pragma: no cover - optional until MongoDB is configured
    ASCENDING = DESCENDING = MongoClient = ReturnDocument = None
    DuplicateKeyError = PyMongoError = Exception


IST = timezone(timedelta(hours=5, minutes=30), name="IST")
DB_ENV_VAR = "DATALENS_OUTSOURCE_ATTENDANCE_DB"
ADMIN_PASSWORD_ENV_VAR = "DATALENS_ATTENDANCE_ADMIN_PASSWORD"
MONGODB_URI_ENV_VAR = "MONGODB_URI"
MONGODB_DATABASE_ENV_VAR = "MONGODB_DATABASE"
DEFAULT_MONGODB_DATABASE = "attendance_db"
DEFAULT_ADMIN_PASSWORD = "admin123"
VALID_USER_ROLES = {"observer", "outsource"}
VALID_DECISIONS = {"accepted", "rejected"}
SHIFT_ORDER = {"M": 1, "G": 2, "E": 3, "N": 4, "O": 5}
SHIFT_LABELS = {
    "M": "Morning",
    "G": "General",
    "E": "Evening",
    "N": "Night",
    "O": "Other",
}


def now_ist() -> datetime:
    """Return the current Indian Standard Time."""
    return datetime.now(IST)


def classify_shift(moment: datetime | time) -> tuple[str, str]:
    """Classify an IST login time into the requested shift code."""
    login_time = moment.timetz() if isinstance(moment, datetime) else moment
    hour = login_time.hour

    if 7 <= hour < 9:
        return "M", SHIFT_LABELS["M"]
    if 9 <= hour <= 12:
        return "G", SHIFT_LABELS["G"]
    if 13 <= hour <= 16:
        return "E", SHIFT_LABELS["E"]
    if 19 <= hour <= 21:
        return "N", SHIFT_LABELS["N"]
    return "O", SHIFT_LABELS["O"]


def get_default_db_path(path: str | os.PathLike[str] | None = None) -> Path:
    """Resolve the local SQLite file used for outsource attendance."""
    if path is not None:
        return Path(path)

    env_path = os.environ.get(DB_ENV_VAR)
    if env_path:
        return Path(env_path)

    return Path(__file__).resolve().parent.parent / "data" / "outsource_attendance.sqlite"


def _clean_name(value: str) -> str:
    return " ".join(str(value or "").strip().split())


def _clean_text(value: str) -> str:
    return " ".join(str(value or "").strip().split())


def _clean_mobile(value: str) -> str:
    return "".join(ch for ch in str(value or "") if ch.isdigit())


def _normalize_name(value: str) -> str:
    return _clean_name(value).casefold()


def _clean_pc_name(value: str) -> str:
    return " ".join(str(value or "").strip().upper().split())


def _coerce_ist(moment: datetime | None) -> datetime:
    if moment is None:
        return now_ist()
    if moment.tzinfo is None:
        return moment.replace(tzinfo=IST)
    return moment.astimezone(IST)


def _timestamp() -> str:
    return now_ist().isoformat(timespec="seconds")


def _month_label(month: str) -> str:
    try:
        parsed = datetime.strptime(month, "%Y-%m")
    except ValueError:
        return month
    return parsed.strftime("%B %Y")


def _get_config_value(key: str, default: str = "") -> str:
    env_value = os.environ.get(key)
    if env_value:
        return env_value
    try:
        value = st.secrets.get(key, default)
    except Exception:
        return default
    return str(value) if value is not None else default


def hash_password(password: str) -> str:
    return hashlib.sha256(str(password or "").encode("utf-8")).hexdigest()


def verify_password(password: str, password_hash: str | None) -> bool:
    if not password_hash:
        return False
    return hmac.compare_digest(hash_password(password), str(password_hash))


def get_admin_password_hash() -> str:
    return hash_password(_get_config_value(ADMIN_PASSWORD_ENV_VAR, DEFAULT_ADMIN_PASSWORD))


@dataclass
class AttendanceService:
    """Small SQLite service for outsource login attendance."""

    db_path: str | os.PathLike[str] | None = None

    def __post_init__(self) -> None:
        self.db_path = get_default_db_path(self.db_path)
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self.storage_label = f"SQLite database: {self.db_path}"
        self.initialize()

    def _connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        return conn

    def initialize(self) -> None:
        """Create database tables if they do not exist."""
        with self._connect() as conn:
            conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS users (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL,
                    normalized_name TEXT NOT NULL,
                    mobile TEXT,
                    normalized_mobile TEXT,
                    password_hash TEXT,
                    role TEXT NOT NULL CHECK (role IN ('observer', 'outsource')),
                    agency TEXT,
                    designation TEXT,
                    details TEXT,
                    active INTEGER NOT NULL DEFAULT 1,
                    created_at TEXT NOT NULL,
                    created_by TEXT,
                    updated_at TEXT NOT NULL
                );

                CREATE UNIQUE INDEX IF NOT EXISTS idx_att_users_role_name
                ON users(role, normalized_name);

                CREATE TABLE IF NOT EXISTS login_entries (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    outsource_user_id INTEGER NOT NULL,
                    outsource_name TEXT NOT NULL,
                    pc_name TEXT NOT NULL,
                    login_time_ist TEXT NOT NULL,
                    login_date TEXT NOT NULL,
                    shift_code TEXT NOT NULL,
                    shift_name TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    observer_status TEXT CHECK (
                        observer_status IN ('accepted', 'rejected') OR observer_status IS NULL
                    ),
                    observer_by TEXT,
                    observer_decided_at TEXT,
                    observer_remarks TEXT,
                    admin_status TEXT CHECK (
                        admin_status IN ('accepted', 'rejected') OR admin_status IS NULL
                    ),
                    admin_by TEXT,
                    admin_decided_at TEXT,
                    admin_remarks TEXT,
                    FOREIGN KEY(outsource_user_id) REFERENCES users(id)
                );

                CREATE INDEX IF NOT EXISTS idx_att_entries_login_date
                ON login_entries(login_date);

                CREATE INDEX IF NOT EXISTS idx_att_entries_user_date
                ON login_entries(outsource_user_id, login_date);

                CREATE TABLE IF NOT EXISTS audit_log (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    event_type TEXT NOT NULL,
                    actor_role TEXT,
                    actor_name TEXT,
                    entry_id INTEGER,
                    user_id INTEGER,
                    details TEXT,
                    created_at TEXT NOT NULL
                );

                CREATE INDEX IF NOT EXISTS idx_att_audit_created_at
                ON audit_log(created_at);
                """
            )
            self._migrate_user_columns(conn)

    def _migrate_user_columns(self, conn: sqlite3.Connection) -> None:
        existing_columns = {
            row["name"]
            for row in conn.execute("PRAGMA table_info(users)").fetchall()
        }
        migrations = {
            "mobile": "ALTER TABLE users ADD COLUMN mobile TEXT",
            "normalized_mobile": "ALTER TABLE users ADD COLUMN normalized_mobile TEXT",
            "password_hash": "ALTER TABLE users ADD COLUMN password_hash TEXT",
            "agency": "ALTER TABLE users ADD COLUMN agency TEXT",
            "designation": "ALTER TABLE users ADD COLUMN designation TEXT",
            "details": "ALTER TABLE users ADD COLUMN details TEXT",
        }
        for column, statement in migrations.items():
            if column not in existing_columns:
                conn.execute(statement)

        conn.execute(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS idx_att_users_role_mobile
            ON users(role, normalized_mobile)
            WHERE normalized_mobile IS NOT NULL AND normalized_mobile <> ''
            """
        )

    def _log_event(
        self,
        conn: sqlite3.Connection,
        event_type: str,
        actor_role: str | None = None,
        actor_name: str | None = None,
        entry_id: int | None = None,
        user_id: int | None = None,
        details: str | None = None,
    ) -> None:
        conn.execute(
            """
            INSERT INTO audit_log (
                event_type, actor_role, actor_name, entry_id, user_id, details, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (event_type, actor_role, actor_name, entry_id, user_id, details, _timestamp()),
        )

    def add_user(
        self,
        name: str,
        role: str,
        mobile: str,
        password: str = "",
        designation: str = "",
        details: str = "",
        created_by: str = "Admin",
    ) -> int:
        """Add an observer or outsource user."""
        role = str(role or "").strip().lower()
        clean_name = _clean_name(name)
        clean_mobile = _clean_mobile(mobile)
        if role not in VALID_USER_ROLES:
            raise ValueError("User role must be observer or outsource.")
        if not clean_name:
            raise ValueError("User name is required.")
        if not clean_mobile:
            raise ValueError("Mobile number is required.")

        created_by = _clean_name(created_by) or "Admin"
        normalized = _normalize_name(clean_name)
        password_hash = None
        if role == "observer":
            password_hash = hash_password(str(password or "").strip() or clean_mobile)
        timestamp = _timestamp()

        with self._connect() as conn:
            try:
                cursor = conn.execute(
                    """
                    INSERT INTO users (
                        name, normalized_name, mobile, normalized_mobile, password_hash,
                        role, agency, designation, details, active, created_at, created_by, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?, ?)
                    """,
                    (
                        clean_name,
                        normalized,
                        clean_mobile,
                        clean_mobile,
                        password_hash,
                        role,
                        "",
                        _clean_text(designation),
                        str(details or "").strip(),
                        timestamp,
                        created_by,
                        timestamp,
                    ),
                )
                user_id = int(cursor.lastrowid)
                self._log_event(
                    conn,
                    "user_created",
                    actor_role="admin",
                    actor_name=created_by,
                    user_id=user_id,
                    details=f"{role} user {clean_name} created with mobile {clean_mobile}",
                )
                return user_id
            except sqlite3.IntegrityError as exc:
                row = conn.execute(
                    """
                    SELECT id, active FROM users
                    WHERE role = ? AND normalized_name = ?
                    """,
                    (role, normalized),
                ).fetchone()
                if row and not bool(row["active"]):
                    conn.execute(
                        """
                        UPDATE users
                        SET active = 1,
                            name = ?,
                            mobile = ?,
                            normalized_mobile = ?,
                            password_hash = ?,
                            agency = ?,
                            designation = ?,
                            details = ?,
                            updated_at = ?
                        WHERE id = ?
                        """,
                        (
                                clean_name,
                                clean_mobile,
                                clean_mobile,
                                password_hash,
                                "",
                                _clean_text(designation),
                            str(details or "").strip(),
                            timestamp,
                            row["id"],
                        ),
                    )
                    self._log_event(
                        conn,
                        "user_reactivated",
                        actor_role="admin",
                        actor_name=created_by,
                        user_id=int(row["id"]),
                        details=f"{role} user {clean_name} reactivated with updated login details",
                    )
                    return int(row["id"])
                raise ValueError(
                    f"Active {role} user already exists with this name or mobile number."
                ) from exc

    def update_user_profile(
        self,
        user_id: int,
        name: str,
        mobile: str,
        role: str,
        designation: str = "",
        details: str = "",
        password: str = "",
        actor_name: str = "Admin",
    ) -> None:
        """Update user details and optionally reset the login password."""
        role = str(role or "").strip().lower()
        clean_name = _clean_name(name)
        clean_mobile = _clean_mobile(mobile)
        if role not in VALID_USER_ROLES:
            raise ValueError("User role must be observer or outsource.")
        if not clean_name:
            raise ValueError("User name is required.")
        if not clean_mobile:
            raise ValueError("Mobile number is required.")

        set_clauses = [
            "name = ?",
            "normalized_name = ?",
            "mobile = ?",
            "normalized_mobile = ?",
            "role = ?",
            "designation = ?",
            "details = ?",
            "updated_at = ?",
        ]
        values: list[Any] = [
            clean_name,
            _normalize_name(clean_name),
            clean_mobile,
            clean_mobile,
            role,
            _clean_text(designation),
            str(details or "").strip(),
            _timestamp(),
        ]
        if role == "observer" and str(password or "").strip():
            set_clauses.insert(4, "password_hash = ?")
            values.insert(4, hash_password(password))
        elif role == "outsource":
            set_clauses.insert(4, "password_hash = NULL")
        values.append(user_id)

        with self._connect() as conn:
            row = conn.execute("SELECT id FROM users WHERE id = ?", (user_id,)).fetchone()
            if row is None:
                raise ValueError("User not found.")
            try:
                conn.execute(
                    f"""
                    UPDATE users
                    SET {", ".join(set_clauses)}
                    WHERE id = ?
                    """,
                    values,
                )
            except sqlite3.IntegrityError as exc:
                raise ValueError("Another user already has this name or mobile number.") from exc

            self._log_event(
                conn,
                "user_profile_updated",
                actor_role="admin",
                actor_name=_clean_name(actor_name) or "Admin",
                user_id=user_id,
                details=f"{role} user {clean_name} profile updated",
            )

    def set_user_active(self, user_id: int, active: bool, actor_name: str = "Admin") -> None:
        timestamp = _timestamp()
        with self._connect() as conn:
            row = conn.execute("SELECT id, name, role FROM users WHERE id = ?", (user_id,)).fetchone()
            if row is None:
                raise ValueError("User not found.")
            conn.execute(
                "UPDATE users SET active = ?, updated_at = ? WHERE id = ?",
                (1 if active else 0, timestamp, user_id),
            )
            action = "user_reactivated" if active else "user_deactivated"
            self._log_event(
                conn,
                action,
                actor_role="admin",
                actor_name=_clean_name(actor_name) or "Admin",
                user_id=user_id,
                details=f"{row['role']} user {row['name']} {'activated' if active else 'deactivated'}",
            )

    def list_users(self, role: str | None = None, active: bool | None = None) -> pd.DataFrame:
        clauses: list[str] = []
        params: list[Any] = []
        if role:
            clauses.append("role = ?")
            params.append(role)
        if active is not None:
            clauses.append("active = ?")
            params.append(1 if active else 0)

        where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
        query = f"""
            SELECT
                id,
                name,
                mobile,
                role,
                designation,
                details,
                active,
                CASE WHEN password_hash IS NULL OR password_hash = '' THEN 0 ELSE 1 END AS has_password,
                created_at,
                created_by,
                updated_at
            FROM users
            {where}
            ORDER BY role, active DESC, name COLLATE NOCASE
        """
        with self._connect() as conn:
            df = pd.read_sql_query(query, conn, params=params)
        if not df.empty:
            df["active"] = df["active"].astype(bool)
            df["has_password"] = df["has_password"].astype(bool)
        return df

    def authenticate_user(
        self,
        user_id: int,
        password: str,
        role: str,
    ) -> dict[str, Any] | None:
        """Return an active user if the password matches."""
        with self._connect() as conn:
            row = conn.execute(
                """
                SELECT id, name, mobile, role, designation, password_hash
                FROM users
                WHERE id = ? AND role = ? AND active = 1
                """,
                (user_id, role),
            ).fetchone()
            if row is None or not verify_password(password, row["password_hash"]):
                return None

            self._log_event(
                conn,
                "user_login",
                actor_role=role,
                actor_name=row["name"],
                user_id=int(row["id"]),
                details=f"{role} login successful",
            )
            return {
                "id": int(row["id"]),
                "name": row["name"],
                "mobile": row["mobile"] or "",
                "role": row["role"],
                "designation": row["designation"] or "",
            }

    def authenticate_outsource_user(
        self,
        user_id: int,
        mobile: str,
    ) -> dict[str, Any] | None:
        """Return an active outsource user when the mobile number matches."""
        clean_mobile = _clean_mobile(mobile)
        with self._connect() as conn:
            row = conn.execute(
                """
                SELECT id, name, mobile, role, designation
                FROM users
                WHERE id = ?
                    AND role = 'outsource'
                    AND active = 1
                    AND normalized_mobile = ?
                """,
                (user_id, clean_mobile),
            ).fetchone()
            if row is None:
                return None

            self._log_event(
                conn,
                "user_login",
                actor_role="outsource",
                actor_name=row["name"],
                user_id=int(row["id"]),
                details="outsource mobile verification successful",
            )
            return {
                "id": int(row["id"]),
                "name": row["name"],
                "mobile": row["mobile"] or "",
                "role": row["role"],
                "designation": row["designation"] or "",
            }

    def submit_login(
        self,
        outsource_user_id: int,
        pc_name: str,
        login_at: datetime | None = None,
    ) -> int:
        """Submit a login entry from the outsource page."""
        clean_pc = _clean_pc_name(pc_name)
        if not clean_pc:
            raise ValueError("PC name is required.")

        login_time = _coerce_ist(login_at)
        shift_code, shift_name = classify_shift(login_time)
        timestamp = _timestamp()

        with self._connect() as conn:
            user = conn.execute(
                """
                SELECT id, name FROM users
                WHERE id = ? AND role = 'outsource' AND active = 1
                """,
                (outsource_user_id,),
            ).fetchone()
            if user is None:
                raise ValueError("Select an active outsource user.")

            cursor = conn.execute(
                """
                INSERT INTO login_entries (
                    outsource_user_id, outsource_name, pc_name, login_time_ist,
                    login_date, shift_code, shift_name, created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    int(user["id"]),
                    user["name"],
                    clean_pc,
                    login_time.isoformat(timespec="seconds"),
                    login_time.date().isoformat(),
                    shift_code,
                    shift_name,
                    timestamp,
                ),
            )
            entry_id = int(cursor.lastrowid)
            self._log_event(
                conn,
                "login_submitted",
                actor_role="outsource",
                actor_name=user["name"],
                entry_id=entry_id,
                user_id=int(user["id"]),
                details=f"{user['name']} submitted login from {clean_pc} as {shift_code}",
            )
            return entry_id

    def decide_entry(
        self,
        entry_id: int,
        decision: str,
        actor_role: str,
        actor_name: str,
        remarks: str = "",
    ) -> None:
        """Accept or reject a login entry as observer or admin."""
        decision = str(decision or "").strip().lower()
        actor_role = str(actor_role or "").strip().lower()
        actor_name = _clean_name(actor_name) or actor_role.title()
        remarks = str(remarks or "").strip()

        if decision not in VALID_DECISIONS:
            raise ValueError("Decision must be accepted or rejected.")
        if actor_role not in {"observer", "admin"}:
            raise ValueError("Actor role must be observer or admin.")

        timestamp = _timestamp()
        with self._connect() as conn:
            row = conn.execute(
                """
                SELECT id, observer_status, admin_status FROM login_entries
                WHERE id = ?
                """,
                (entry_id,),
            ).fetchone()
            if row is None:
                raise ValueError("Entry not found.")
            if actor_role == "observer" and (row["observer_status"] or row["admin_status"]):
                raise ValueError("Observer can decide only pending entries.")

            if actor_role == "admin":
                conn.execute(
                    """
                    UPDATE login_entries
                    SET admin_status = ?, admin_by = ?, admin_decided_at = ?, admin_remarks = ?
                    WHERE id = ?
                    """,
                    (decision, actor_name, timestamp, remarks, entry_id),
                )
            else:
                conn.execute(
                    """
                    UPDATE login_entries
                    SET observer_status = ?, observer_by = ?, observer_decided_at = ?,
                        observer_remarks = ?
                    WHERE id = ?
                    """,
                    (decision, actor_name, timestamp, remarks, entry_id),
                )

            self._log_event(
                conn,
                f"{actor_role}_{decision}",
                actor_role=actor_role,
                actor_name=actor_name,
                entry_id=entry_id,
                details=remarks or None,
            )

    def list_entries(
        self,
        month: str | None = None,
        status_filter: str | None = None,
        outsource_user_id: int | None = None,
    ) -> pd.DataFrame:
        clauses: list[str] = []
        params: list[Any] = []
        if month:
            clauses.append("substr(e.login_date, 1, 7) = ?")
            params.append(month)
        if outsource_user_id:
            clauses.append("e.outsource_user_id = ?")
            params.append(outsource_user_id)

        where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
        query = f"""
            SELECT
                e.id,
                e.outsource_user_id,
                e.outsource_name,
                COALESCE(u.active, 0) AS outsource_active,
                e.pc_name,
                e.login_time_ist,
                e.login_date,
                e.shift_code,
                e.shift_name,
                e.created_at,
                e.observer_status,
                e.observer_by,
                e.observer_decided_at,
                e.observer_remarks,
                e.admin_status,
                e.admin_by,
                e.admin_decided_at,
                e.admin_remarks
            FROM login_entries e
            LEFT JOIN users u ON u.id = e.outsource_user_id
            {where}
            ORDER BY e.login_time_ist DESC, e.id DESC
        """
        with self._connect() as conn:
            df = pd.read_sql_query(query, conn, params=params)

        if df.empty:
            return self._with_status_columns(df)

        df = self._with_status_columns(df)
        if status_filter and status_filter.lower() != "all":
            df = df[df["effective_status"] == status_filter.lower()].copy()
        return df.reset_index(drop=True)

    def _with_status_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            for column in ["effective_status", "decision_source", "decided_by", "final_remarks"]:
                df[column] = []
            return df

        df = df.copy()
        df["effective_status"] = df["admin_status"].fillna(df["observer_status"]).fillna("pending")
        df["decision_source"] = "Pending"
        df.loc[df["observer_status"].notna(), "decision_source"] = "Observer"
        df.loc[df["admin_status"].notna(), "decision_source"] = "Admin Override"
        df["decided_by"] = df["admin_by"].fillna(df["observer_by"]).fillna("")
        df["final_remarks"] = df["admin_remarks"].fillna(df["observer_remarks"]).fillna("")
        return df

    def get_available_months(self) -> list[str]:
        with self._connect() as conn:
            rows = conn.execute(
                """
                SELECT DISTINCT substr(login_date, 1, 7) AS month
                FROM login_entries
                ORDER BY month DESC
                """
            ).fetchall()
        return [row["month"] for row in rows if row["month"]]

    def get_summary_metrics(self) -> dict[str, int]:
        entries = self.list_entries()
        users = self.list_users()
        today = now_ist().date().isoformat()
        if entries.empty:
            return {
                "pending": 0,
                "accepted": 0,
                "rejected": 0,
                "today_logins": 0,
                "observer_users": int((users["role"] == "observer").sum()) if not users.empty else 0,
                "outsource_users": int((users["role"] == "outsource").sum()) if not users.empty else 0,
            }

        status_counts = entries["effective_status"].value_counts()
        return {
            "pending": int(status_counts.get("pending", 0)),
            "accepted": int(status_counts.get("accepted", 0)),
            "rejected": int(status_counts.get("rejected", 0)),
            "today_logins": int((entries["login_date"] == today).sum()),
            "observer_users": int((users["role"] == "observer").sum()) if not users.empty else 0,
            "outsource_users": int((users["role"] == "outsource").sum()) if not users.empty else 0,
        }

    def build_raw_attendance_df(self, month: str | None = None) -> pd.DataFrame:
        entries = self.list_entries(month=month)
        if entries.empty:
            return pd.DataFrame(
                columns=[
                    "Entry ID",
                    "Outsource Name",
                    "PC Name",
                    "Login Time (IST)",
                    "Login Date",
                    "Shift",
                    "Observer Status",
                    "Observer By",
                    "Observer Remarks",
                    "Admin Status",
                    "Admin By",
                    "Admin Remarks",
                    "Final Status",
                    "Decision Source",
                    "Final Remarks",
                ]
            )

        raw = entries.copy()
        raw["Shift"] = raw["shift_code"] + " - " + raw["shift_name"]
        raw = raw.rename(
            columns={
                "id": "Entry ID",
                "outsource_name": "Outsource Name",
                "pc_name": "PC Name",
                "login_time_ist": "Login Time (IST)",
                "login_date": "Login Date",
                "observer_status": "Observer Status",
                "observer_by": "Observer By",
                "observer_remarks": "Observer Remarks",
                "admin_status": "Admin Status",
                "admin_by": "Admin By",
                "admin_remarks": "Admin Remarks",
                "effective_status": "Final Status",
                "decision_source": "Decision Source",
                "final_remarks": "Final Remarks",
            }
        )
        return raw[
            [
                "Entry ID",
                "Outsource Name",
                "PC Name",
                "Login Time (IST)",
                "Login Date",
                "Shift",
                "Observer Status",
                "Observer By",
                "Observer Remarks",
                "Admin Status",
                "Admin By",
                "Admin Remarks",
                "Final Status",
                "Decision Source",
                "Final Remarks",
            ]
        ]

    def build_monthly_attendance_df(self, month: str) -> pd.DataFrame:
        parsed = datetime.strptime(month, "%Y-%m")
        days_in_month = calendar.monthrange(parsed.year, parsed.month)[1]
        day_columns = [
            f"{day:02d} {date(parsed.year, parsed.month, day).strftime('%a')}"
            for day in range(1, days_in_month + 1)
        ]

        entries = self.list_entries(month=month)
        users = self.list_users(role="outsource")
        names: list[str] = []
        if not users.empty:
            names.extend(users["name"].tolist())
        if not entries.empty:
            names.extend(entries["outsource_name"].dropna().tolist())
        unique_names = sorted(dict.fromkeys(names), key=str.casefold)

        rows: list[dict[str, Any]] = []
        for name in unique_names:
            row: dict[str, Any] = {"Name": name}
            present_days = 0
            employee_entries = (
                entries[entries["outsource_name"] == name].copy()
                if not entries.empty
                else pd.DataFrame()
            )

            for day, column in enumerate(day_columns, start=1):
                current_date = date(parsed.year, parsed.month, day).isoformat()
                marker = ""
                if not employee_entries.empty:
                    day_entries = employee_entries[employee_entries["login_date"] == current_date]
                    accepted = day_entries[day_entries["effective_status"] == "accepted"]
                    pending = day_entries[day_entries["effective_status"] == "pending"]
                    if not accepted.empty:
                        shifts = sorted(
                            accepted["shift_code"].dropna().unique().tolist(),
                            key=lambda code: SHIFT_ORDER.get(code, 99),
                        )
                        marker = "/".join(shifts)
                        present_days += 1
                    elif not pending.empty:
                        marker = "P"
                row[column] = marker
            row["Total Present Days"] = present_days
            rows.append(row)

        return pd.DataFrame(rows, columns=["Name", *day_columns, "Total Present Days"])

    def build_daily_summary_df(self, month: str) -> pd.DataFrame:
        parsed = datetime.strptime(month, "%Y-%m")
        days_in_month = calendar.monthrange(parsed.year, parsed.month)[1]
        entries = self.list_entries(month=month)
        rows: list[dict[str, Any]] = []

        for day in range(1, days_in_month + 1):
            current = date(parsed.year, parsed.month, day)
            current_entries = (
                entries[entries["login_date"] == current.isoformat()]
                if not entries.empty
                else pd.DataFrame()
            )
            accepted = (
                current_entries[current_entries["effective_status"] == "accepted"]
                if not current_entries.empty
                else pd.DataFrame()
            )
            rows.append(
                {
                    "Date": current.isoformat(),
                    "Day": current.strftime("%A"),
                    "Total Logins": int(len(current_entries)),
                    "Accepted": int((current_entries["effective_status"] == "accepted").sum())
                    if not current_entries.empty
                    else 0,
                    "Rejected": int((current_entries["effective_status"] == "rejected").sum())
                    if not current_entries.empty
                    else 0,
                    "Pending": int((current_entries["effective_status"] == "pending").sum())
                    if not current_entries.empty
                    else 0,
                    "Morning": int((accepted["shift_code"] == "M").sum()) if not accepted.empty else 0,
                    "General": int((accepted["shift_code"] == "G").sum()) if not accepted.empty else 0,
                    "Evening": int((accepted["shift_code"] == "E").sum()) if not accepted.empty else 0,
                    "Night": int((accepted["shift_code"] == "N").sum()) if not accepted.empty else 0,
                    "Other": int((accepted["shift_code"] == "O").sum()) if not accepted.empty else 0,
                }
            )

        return pd.DataFrame(rows)

    def get_audit_log(self, limit: int = 500) -> pd.DataFrame:
        with self._connect() as conn:
            return pd.read_sql_query(
                """
                SELECT
                    id AS "Audit ID",
                    event_type AS "Event",
                    actor_role AS "Actor Role",
                    actor_name AS "Actor Name",
                    entry_id AS "Entry ID",
                    user_id AS "User ID",
                    details AS "Details",
                    created_at AS "Created At"
                FROM audit_log
                ORDER BY created_at DESC, id DESC
                LIMIT ?
                """,
                conn,
                params=[limit],
            )

    def export_attendance_workbook(self, month: str) -> bytes:
        output = BytesIO()
        sheets = {
            "Login Register": self.build_raw_attendance_df(month),
            "Monthly Attendance": self.build_monthly_attendance_df(month),
            "Daily Summary": self.build_daily_summary_df(month),
            "User Master": self.list_users(),
            "Audit Log": self.get_audit_log(limit=5000),
        }

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            for sheet_name, df in sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
            self._style_workbook(writer.book, month)

        output.seek(0)
        return output.getvalue()

    def _style_workbook(self, workbook: Any, month: str) -> None:
        title_fill = PatternFill("solid", fgColor="111827")
        subtitle_fill = PatternFill("solid", fgColor="E8EEF7")
        header_fill = PatternFill("solid", fgColor="C9DAF8")
        border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )
        status_fill = {
            "accepted": PatternFill("solid", fgColor="D9EAD3"),
            "rejected": PatternFill("solid", fgColor="F4CCCC"),
            "pending": PatternFill("solid", fgColor="FFF2CC"),
            "P": PatternFill("solid", fgColor="FFF2CC"),
            "R": PatternFill("solid", fgColor="F4CCCC"),
        }
        shift_fill = {
            "M": PatternFill("solid", fgColor="D9EAD3"),
            "G": PatternFill("solid", fgColor="D9EAF7"),
            "E": PatternFill("solid", fgColor="FCE5CD"),
            "N": PatternFill("solid", fgColor="D9D2E9"),
            "O": PatternFill("solid", fgColor="EADCF8"),
        }

        for worksheet in workbook.worksheets:
            max_column = max(worksheet.max_column, 1)
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_column)
            worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_column)
            worksheet["A1"] = f"{worksheet.title} - Outsource Attendance"
            subtitle = (
                f"Month: {_month_label(month)} | Generated: "
                f"{now_ist().strftime('%d-%m-%Y %H:%M IST')}"
            )
            if worksheet.title == "Monthly Attendance":
                subtitle += " | M 07-08, G 09-12, E 13-16, N 19-21, O other"
            worksheet["A2"] = subtitle
            worksheet["A1"].fill = title_fill
            worksheet["A1"].font = Font(color="FFFFFF", bold=True, size=15)
            worksheet["A1"].alignment = Alignment(horizontal="center")
            worksheet["A2"].fill = subtitle_fill
            worksheet["A2"].font = Font(color="334155", italic=True)
            worksheet["A2"].alignment = Alignment(horizontal="center")

            for cell in worksheet[3]:
                cell.fill = header_fill
                cell.font = Font(bold=True, color="111827")
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for row in worksheet.iter_rows(min_row=4):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    value = str(cell.value or "")
                    lower_value = value.lower()
                    if lower_value in status_fill:
                        cell.fill = status_fill[lower_value]
                    elif value in status_fill:
                        cell.fill = status_fill[value]
                    elif value in shift_fill:
                        cell.fill = shift_fill[value]
                    elif "/" in value and all(part in shift_fill for part in value.split("/")):
                        cell.fill = PatternFill("solid", fgColor="E2F0D9")

            for index, column_cells in enumerate(worksheet.columns, start=1):
                values = [str(cell.value or "") for cell in column_cells]
                width = min(max(max((len(value) for value in values), default=8) + 2, 10), 34)
                if worksheet.title == "Monthly Attendance" and index > 1:
                    width = 11 if index < worksheet.max_column else 18
                worksheet.column_dimensions[get_column_letter(index)].width = width

            worksheet.freeze_panes = "A4"
            worksheet.sheet_view.showGridLines = False
            if worksheet.max_row >= 3:
                worksheet.auto_filter.ref = f"A3:{get_column_letter(max_column)}{worksheet.max_row}"


class MongoAttendanceService(AttendanceService):
    """MongoDB Atlas-backed attendance service with the SQLite service API."""

    def __init__(self, uri: str, database_name: str = DEFAULT_MONGODB_DATABASE) -> None:
        if MongoClient is None:
            raise RuntimeError("pymongo is required for MongoDB attendance storage.")
        self.uri = uri
        self.database_name = database_name or DEFAULT_MONGODB_DATABASE
        self.client = MongoClient(uri, serverSelectionTimeoutMS=8000)
        self.db = self.client[self.database_name]
        self.users = self.db["attendance_users"]
        self.entries = self.db["attendance_login_entries"]
        self.audit = self.db["attendance_audit_log"]
        self.counters = self.db["attendance_counters"]
        self.storage_label = f"MongoDB database: {self.database_name}"
        self.initialize()

    def initialize(self) -> None:
        self.client.admin.command("ping")
        self.users.create_index(
            [("role", ASCENDING), ("normalized_name", ASCENDING)],
            unique=True,
            name="idx_att_users_role_name",
        )
        self.users.create_index(
            [("role", ASCENDING), ("normalized_mobile", ASCENDING)],
            unique=True,
            sparse=True,
            name="idx_att_users_role_mobile",
        )
        self.entries.create_index([("login_date", ASCENDING)], name="idx_att_entries_login_date")
        self.entries.create_index(
            [("outsource_user_id", ASCENDING), ("login_date", ASCENDING)],
            name="idx_att_entries_user_date",
        )
        self.audit.create_index([("created_at", DESCENDING)], name="idx_att_audit_created_at")

    def _next_id(self, name: str) -> int:
        row = self.counters.find_one_and_update(
            {"_id": name},
            {"$inc": {"seq": 1}},
            upsert=True,
            return_document=ReturnDocument.AFTER,
        )
        return int(row["seq"])

    def _log_event(
        self,
        _conn: Any,
        event_type: str,
        actor_role: str | None = None,
        actor_name: str | None = None,
        entry_id: int | None = None,
        user_id: int | None = None,
        details: str | None = None,
    ) -> None:
        self.audit.insert_one(
            {
                "id": self._next_id("audit_log"),
                "event_type": event_type,
                "actor_role": actor_role,
                "actor_name": actor_name,
                "entry_id": entry_id,
                "user_id": user_id,
                "details": details,
                "created_at": _timestamp(),
            }
        )

    @staticmethod
    def _docs_to_df(docs: list[dict[str, Any]]) -> pd.DataFrame:
        df = pd.DataFrame(docs)
        if "_id" in df.columns:
            df = df.drop(columns=["_id"])
        return df

    def add_user(
        self,
        name: str,
        role: str,
        mobile: str,
        password: str = "",
        designation: str = "",
        details: str = "",
        created_by: str = "Admin",
    ) -> int:
        role = str(role or "").strip().lower()
        clean_name = _clean_name(name)
        clean_mobile = _clean_mobile(mobile)
        if role not in VALID_USER_ROLES:
            raise ValueError("User role must be observer or outsource.")
        if not clean_name:
            raise ValueError("User name is required.")
        if not clean_mobile:
            raise ValueError("Mobile number is required.")

        timestamp = _timestamp()
        password_hash = hash_password(str(password or "").strip() or clean_mobile) if role == "observer" else None
        doc = {
            "id": self._next_id("users"),
            "name": clean_name,
            "normalized_name": _normalize_name(clean_name),
            "mobile": clean_mobile,
            "normalized_mobile": clean_mobile,
            "password_hash": password_hash,
            "role": role,
            "designation": _clean_text(designation),
            "details": str(details or "").strip(),
            "active": True,
            "created_at": timestamp,
            "created_by": _clean_name(created_by) or "Admin",
            "updated_at": timestamp,
        }
        try:
            self.users.insert_one(doc)
        except DuplicateKeyError as exc:
            raise ValueError("Active user already exists with this name or mobile number.") from exc

        self._log_event(
            None,
            "user_created",
            actor_role="admin",
            actor_name=doc["created_by"],
            user_id=doc["id"],
            details=f"{role} user {clean_name} created with mobile {clean_mobile}",
        )
        return int(doc["id"])

    def update_user_profile(
        self,
        user_id: int,
        name: str,
        mobile: str,
        role: str,
        designation: str = "",
        details: str = "",
        password: str = "",
        actor_name: str = "Admin",
    ) -> None:
        role = str(role or "").strip().lower()
        clean_name = _clean_name(name)
        clean_mobile = _clean_mobile(mobile)
        if role not in VALID_USER_ROLES:
            raise ValueError("User role must be observer or outsource.")
        if not clean_name:
            raise ValueError("User name is required.")
        if not clean_mobile:
            raise ValueError("Mobile number is required.")

        update_doc: dict[str, Any] = {
            "name": clean_name,
            "normalized_name": _normalize_name(clean_name),
            "mobile": clean_mobile,
            "normalized_mobile": clean_mobile,
            "role": role,
            "designation": _clean_text(designation),
            "details": str(details or "").strip(),
            "updated_at": _timestamp(),
        }
        if role == "observer" and str(password or "").strip():
            update_doc["password_hash"] = hash_password(password)
        if role == "outsource":
            update_doc["password_hash"] = None

        try:
            result = self.users.update_one({"id": int(user_id)}, {"$set": update_doc})
        except DuplicateKeyError as exc:
            raise ValueError("Another user already has this name or mobile number.") from exc
        if result.matched_count == 0:
            raise ValueError("User not found.")

        self._log_event(
            None,
            "user_profile_updated",
            actor_role="admin",
            actor_name=_clean_name(actor_name) or "Admin",
            user_id=int(user_id),
            details=f"{role} user {clean_name} profile updated",
        )

    def set_user_active(self, user_id: int, active: bool, actor_name: str = "Admin") -> None:
        row = self.users.find_one({"id": int(user_id)})
        if row is None:
            raise ValueError("User not found.")
        self.users.update_one(
            {"id": int(user_id)},
            {"$set": {"active": bool(active), "updated_at": _timestamp()}},
        )
        self._log_event(
            None,
            "user_reactivated" if active else "user_deactivated",
            actor_role="admin",
            actor_name=_clean_name(actor_name) or "Admin",
            user_id=int(user_id),
            details=f"{row['role']} user {row['name']} {'activated' if active else 'deactivated'}",
        )

    def list_users(self, role: str | None = None, active: bool | None = None) -> pd.DataFrame:
        query: dict[str, Any] = {}
        if role:
            query["role"] = role
        if active is not None:
            query["active"] = bool(active)
        docs = list(self.users.find(query).sort([("role", ASCENDING), ("active", DESCENDING), ("name", ASCENDING)]))
        df = self._docs_to_df(docs)
        columns = [
            "id", "name", "mobile", "role", "designation", "details", "active",
            "has_password", "created_at", "created_by", "updated_at",
        ]
        if df.empty:
            return pd.DataFrame(columns=columns)
        df["active"] = df["active"].astype(bool)
        df["has_password"] = df.get("password_hash", pd.Series([None] * len(df))).fillna("").astype(bool)
        return df.reindex(columns=columns)

    def authenticate_user(self, user_id: int, password: str, role: str) -> dict[str, Any] | None:
        row = self.users.find_one({"id": int(user_id), "role": role, "active": True})
        if row is None or not verify_password(password, row.get("password_hash")):
            return None
        self._log_event(
            None,
            "user_login",
            actor_role=role,
            actor_name=row["name"],
            user_id=int(row["id"]),
            details=f"{role} login successful",
        )
        return {
            "id": int(row["id"]),
            "name": row["name"],
            "mobile": row.get("mobile", ""),
            "role": row["role"],
            "designation": row.get("designation", ""),
        }

    def authenticate_outsource_user(self, user_id: int, mobile: str) -> dict[str, Any] | None:
        row = self.users.find_one(
            {
                "id": int(user_id),
                "role": "outsource",
                "active": True,
                "normalized_mobile": _clean_mobile(mobile),
            }
        )
        if row is None:
            return None
        self._log_event(
            None,
            "user_login",
            actor_role="outsource",
            actor_name=row["name"],
            user_id=int(row["id"]),
            details="outsource mobile verification successful",
        )
        return {
            "id": int(row["id"]),
            "name": row["name"],
            "mobile": row.get("mobile", ""),
            "role": row["role"],
            "designation": row.get("designation", ""),
        }

    def submit_login(
        self,
        outsource_user_id: int,
        pc_name: str,
        login_at: datetime | None = None,
    ) -> int:
        clean_pc = _clean_pc_name(pc_name)
        if not clean_pc:
            raise ValueError("PC name is required.")
        user = self.users.find_one({"id": int(outsource_user_id), "role": "outsource", "active": True})
        if user is None:
            raise ValueError("Select an active outsource user.")

        login_time = _coerce_ist(login_at)
        shift_code, shift_name = classify_shift(login_time)
        entry_id = self._next_id("login_entries")
        self.entries.insert_one(
            {
                "id": entry_id,
                "outsource_user_id": int(user["id"]),
                "outsource_name": user["name"],
                "pc_name": clean_pc,
                "login_time_ist": login_time.isoformat(timespec="seconds"),
                "login_date": login_time.date().isoformat(),
                "shift_code": shift_code,
                "shift_name": shift_name,
                "created_at": _timestamp(),
                "observer_status": None,
                "observer_by": None,
                "observer_decided_at": None,
                "observer_remarks": None,
                "admin_status": None,
                "admin_by": None,
                "admin_decided_at": None,
                "admin_remarks": None,
            }
        )
        self._log_event(
            None,
            "login_submitted",
            actor_role="outsource",
            actor_name=user["name"],
            entry_id=entry_id,
            user_id=int(user["id"]),
            details=f"{user['name']} submitted login from {clean_pc} as {shift_code}",
        )
        return entry_id

    def decide_entry(
        self,
        entry_id: int,
        decision: str,
        actor_role: str,
        actor_name: str,
        remarks: str = "",
    ) -> None:
        decision = str(decision or "").strip().lower()
        actor_role = str(actor_role or "").strip().lower()
        actor_name = _clean_name(actor_name) or actor_role.title()
        remarks = str(remarks or "").strip()
        if decision not in VALID_DECISIONS:
            raise ValueError("Decision must be accepted or rejected.")
        if actor_role not in {"observer", "admin"}:
            raise ValueError("Actor role must be observer or admin.")

        row = self.entries.find_one({"id": int(entry_id)})
        if row is None:
            raise ValueError("Entry not found.")
        if actor_role == "observer" and (row.get("observer_status") or row.get("admin_status")):
            raise ValueError("Observer can decide only pending entries.")

        timestamp = _timestamp()
        if actor_role == "admin":
            update_doc = {
                "admin_status": decision,
                "admin_by": actor_name,
                "admin_decided_at": timestamp,
                "admin_remarks": remarks,
            }
        else:
            update_doc = {
                "observer_status": decision,
                "observer_by": actor_name,
                "observer_decided_at": timestamp,
                "observer_remarks": remarks,
            }
        self.entries.update_one({"id": int(entry_id)}, {"$set": update_doc})
        self._log_event(
            None,
            f"{actor_role}_{decision}",
            actor_role=actor_role,
            actor_name=actor_name,
            entry_id=int(entry_id),
            details=remarks or None,
        )

    def list_entries(
        self,
        month: str | None = None,
        status_filter: str | None = None,
        outsource_user_id: int | None = None,
    ) -> pd.DataFrame:
        query: dict[str, Any] = {}
        if month:
            query["login_date"] = {"$regex": f"^{month}"}
        if outsource_user_id:
            query["outsource_user_id"] = int(outsource_user_id)
        docs = list(self.entries.find(query).sort([("login_time_ist", DESCENDING), ("id", DESCENDING)]))
        df = self._docs_to_df(docs)
        df = self._with_status_columns(df)
        if not df.empty and status_filter and status_filter.lower() != "all":
            df = df[df["effective_status"] == status_filter.lower()].copy()
        return df.reset_index(drop=True)

    def get_available_months(self) -> list[str]:
        dates = [value for value in self.entries.distinct("login_date") if value]
        return sorted({str(value)[:7] for value in dates}, reverse=True)

    def get_audit_log(self, limit: int = 500) -> pd.DataFrame:
        docs = list(self.audit.find({}).sort([("created_at", DESCENDING), ("id", DESCENDING)]).limit(limit))
        df = self._docs_to_df(docs)
        if df.empty:
            return pd.DataFrame(
                columns=["Audit ID", "Event", "Actor Role", "Actor Name", "Entry ID", "User ID", "Details", "Created At"]
            )
        return df.rename(
            columns={
                "id": "Audit ID",
                "event_type": "Event",
                "actor_role": "Actor Role",
                "actor_name": "Actor Name",
                "entry_id": "Entry ID",
                "user_id": "User ID",
                "details": "Details",
                "created_at": "Created At",
            }
        ).reindex(
            columns=["Audit ID", "Event", "Actor Role", "Actor Name", "Entry ID", "User ID", "Details", "Created At"]
        )


def get_attendance_service() -> AttendanceService:
    """Return MongoDB storage when configured, otherwise local SQLite."""
    mongodb_uri = _get_config_value(MONGODB_URI_ENV_VAR)
    if mongodb_uri:
        database_name = _get_config_value(MONGODB_DATABASE_ENV_VAR, DEFAULT_MONGODB_DATABASE)
        return MongoAttendanceService(mongodb_uri, database_name)
    return AttendanceService()


def _month_options(service: AttendanceService) -> list[str]:
    current = now_ist().strftime("%Y-%m")
    months = service.get_available_months()
    if current not in months:
        months.insert(0, current)
    return months


def _status_title(value: Any) -> str:
    if pd.isna(value) or value in ("", None):
        return ""
    return str(value).replace("_", " ").title()


def _display_entries(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return _empty_entries_display()

    display = df.copy()
    display["Shift"] = display["shift_code"] + " - " + display["shift_name"]
    display = display.rename(
        columns={
            "id": "Entry ID",
            "outsource_name": "Outsource Name",
            "pc_name": "PC Name",
            "login_time_ist": "Login Time (IST)",
            "login_date": "Login Date",
            "observer_status": "Observer Status",
            "observer_by": "Observer By",
            "observer_remarks": "Observer Remarks",
            "admin_status": "Admin Override",
            "admin_by": "Admin By",
            "admin_remarks": "Admin Remarks",
            "effective_status": "Final Status",
            "decision_source": "Decision Source",
            "final_remarks": "Final Remarks",
        }
    )
    columns = [
        "Entry ID",
        "Outsource Name",
        "PC Name",
        "Login Time (IST)",
        "Login Date",
        "Shift",
        "Final Status",
        "Decision Source",
        "Observer Status",
        "Observer By",
        "Observer Remarks",
        "Admin Override",
        "Admin By",
        "Admin Remarks",
        "Final Remarks",
    ]
    for column in ["Observer Status", "Admin Override", "Final Status", "Decision Source"]:
        display[column] = display[column].map(_status_title)
    display["Admin Override"] = display["Admin Override"].replace("", "No override")
    return display[columns].fillna("")


def _empty_entries_display() -> pd.DataFrame:
    return pd.DataFrame(
        columns=[
            "Entry ID",
            "Outsource Name",
            "PC Name",
            "Login Time (IST)",
            "Login Date",
            "Shift",
            "Final Status",
            "Decision Source",
            "Observer Status",
            "Observer By",
            "Observer Remarks",
            "Admin Override",
            "Admin By",
            "Admin Remarks",
            "Final Remarks",
        ]
    )


def _entry_options(df: pd.DataFrame) -> dict[str, int]:
    options: dict[str, int] = {}
    for _, row in df.iterrows():
        label = (
            f"#{row['id']} | {row['outsource_name']} | {row['pc_name']} | "
            f"{row['login_time_ist']} | {_status_title(row['effective_status'])}"
        )
        options[label] = int(row["id"])
    return options


def _observer_actionable_entries(entries: pd.DataFrame) -> pd.DataFrame:
    """Return entries an observer is still allowed to decide."""
    if entries.empty:
        return entries
    return entries[
        entries["admin_status"].isna()
        & entries["observer_status"].isna()
        & (entries["effective_status"] == "pending")
    ].copy()


def _render_metrics(metrics: dict[str, int]) -> None:
    columns = st.columns(6)
    columns[0].metric("Pending", metrics["pending"])
    columns[1].metric("Accepted", metrics["accepted"])
    columns[2].metric("Rejected", metrics["rejected"])
    columns[3].metric("Today Logins", metrics["today_logins"])
    columns[4].metric("Observers", metrics["observer_users"])
    columns[5].metric("Outsource", metrics["outsource_users"])


def _auth_key(role: str) -> str:
    return f"attendance_auth_{role}"


def _logout_authenticated_role(role: str) -> None:
    st.session_state.pop(_auth_key(role), None)
    st.rerun()


def _render_auth_status(auth: dict[str, Any], role: str) -> None:
    status_col, logout_col = st.columns([0.78, 0.22])
    with status_col:
        st.caption(
            f"Logged in as {auth.get('name', role.title())} "
            f"({role.replace('_', ' ').title()})"
        )
    with logout_col:
        if st.button("Logout", use_container_width=True, key=f"{role}_logout"):
            _logout_authenticated_role(role)


def _require_admin_auth() -> dict[str, Any] | None:
    session_key = _auth_key("admin")
    auth = st.session_state.get(session_key)
    if auth:
        _render_auth_status(auth, "admin")
        return auth

    st.title("Attendance Admin Login")
    st.warning("Admin approval, user creation, override, and export are password protected.")
    with st.form("attendance_admin_login_form"):
        admin_name = st.text_input("Admin name", value="Admin")
        password = st.text_input("Admin password", type="password")
        submitted = st.form_submit_button("Login", type="primary", use_container_width=True)
        if submitted:
            if verify_password(password, get_admin_password_hash()):
                st.session_state[session_key] = {
                    "role": "admin",
                    "name": _clean_name(admin_name) or "Admin",
                }
                st.rerun()
            else:
                st.error("Incorrect admin password.")
    st.caption(f"Admin password can be changed with environment variable {ADMIN_PASSWORD_ENV_VAR}.")
    return None


def _require_user_auth(service: AttendanceService, role: str) -> dict[str, Any] | None:
    session_key = _auth_key(role)
    auth = st.session_state.get(session_key)
    if auth:
        _render_auth_status(auth, role)
        return auth

    role_title = role.title()
    st.title(f"{role_title} Login")
    users = service.list_users(role=role, active=True)
    if users.empty:
        st.warning(f"Ask admin to create an active {role} user first.")
        return None

    user_options = {str(row["name"]): int(row["id"]) for _, row in users.iterrows()}
    with st.form(f"attendance_{role}_login_form"):
        selected_label = st.selectbox("User", options=list(user_options.keys()))
        if role == "outsource":
            credential = st.text_input("Mobile number")
            help_text = "Enter your mobile number saved by admin."
        else:
            credential = st.text_input("Password", type="password")
            help_text = "Enter the observer password saved by admin."
        st.caption(help_text)
        submitted = st.form_submit_button("Login", type="primary", use_container_width=True)
        if submitted:
            if role == "outsource":
                user = service.authenticate_outsource_user(
                    user_id=user_options[selected_label],
                    mobile=credential,
                )
            else:
                user = service.authenticate_user(
                    user_id=user_options[selected_label],
                    password=credential,
                    role=role,
                )
            if user:
                st.session_state[session_key] = user
                st.rerun()
            else:
                st.error("Incorrect login detail or inactive user. Ask admin to check the account if needed.")
    return None


def _render_decision_form(
    service: AttendanceService,
    entries: pd.DataFrame,
    actor_role: str,
    actor_name: str,
    key_prefix: str,
) -> None:
    if entries.empty:
        st.info("No entries available for decision.")
        return

    options = _entry_options(entries)
    selected_label = st.selectbox("Select entry", options=list(options.keys()), key=f"{key_prefix}_entry")
    decision_label = st.radio(
        "Decision",
        options=["Accepted", "Rejected"],
        horizontal=True,
        key=f"{key_prefix}_decision",
    )
    remarks = st.text_area(
        "Remarks",
        placeholder="Optional. Add reason when rejecting or context for audit.",
        key=f"{key_prefix}_remarks",
    )
    if st.button("Save Decision", type="primary", use_container_width=True, key=f"{key_prefix}_save"):
        try:
            service.decide_entry(
                entry_id=options[selected_label],
                decision=decision_label.lower(),
                actor_role=actor_role,
                actor_name=actor_name,
                remarks=remarks,
            )
            st.success("Decision saved.")
            st.rerun()
        except ValueError as exc:
            st.error(str(exc))


def render_attendance_admin_page() -> None:
    service = get_attendance_service()
    auth = _require_admin_auth()
    if not auth:
        return

    st.title("Outsource Attendance Admin")
    st.caption(getattr(service, "storage_label", "Attendance storage ready"))

    _render_metrics(service.get_summary_metrics())
    st.markdown("---")

    tabs = st.tabs(["Approvals", "Users", "Attendance", "Export", "Audit"])

    with tabs[0]:
        st.subheader("Admin Decisions")
        filter_col1, filter_col2, filter_col3 = st.columns(3)
        month = filter_col1.selectbox(
            "Month",
            options=_month_options(service),
            format_func=_month_label,
            key="admin_entries_month",
        )
        status = filter_col2.selectbox(
            "Current status",
            options=["all", "pending", "accepted", "rejected"],
            format_func=_status_title,
            key="admin_entries_status",
        )
        outsource_users = service.list_users(role="outsource")
        outsource_options = {"All": None}
        if not outsource_users.empty:
            outsource_options.update(
                {row["name"]: int(row["id"]) for _, row in outsource_users.iterrows()}
            )
        outsource_name = filter_col3.selectbox(
            "Outsource user",
            options=list(outsource_options.keys()),
            key="admin_entries_user",
        )
        entries = service.list_entries(
            month=month,
            status_filter=status,
            outsource_user_id=outsource_options[outsource_name],
        )
        st.dataframe(
            _display_entries(entries) if not entries.empty else _empty_entries_display(),
            use_container_width=True,
            hide_index=True,
        )

        st.markdown("---")
        st.caption(f"Decision by: {auth['name']}")
        st.caption("Use Current status = All to override accepted or rejected entries anytime.")
        action_entries = entries
        _render_decision_form(
            service=service,
            entries=action_entries,
            actor_role="admin",
            actor_name=auth["name"],
            key_prefix="admin_decision",
        )

    with tabs[1]:
        st.subheader("User Management")
        create_col, list_col = st.columns([0.35, 0.65])
        with create_col:
            role = st.radio(
                "User type",
                options=["observer", "outsource"],
                format_func=_status_title,
                horizontal=True,
                key="create_attendance_user_role",
            )
            with st.form("create_attendance_user"):
                name = st.text_input("Name")
                mobile = st.text_input("Mobile number")
                designation = st.text_input("Designation / Work detail")
                if role == "observer":
                    password = st.text_input(
                        "Observer password",
                        type="password",
                        help="Leave blank to use the mobile number as the first password.",
                    )
                else:
                    password = ""
                    st.caption("Outsource users verify login with their registered mobile number.")
                details = st.text_area("Other details", placeholder="Optional notes for admin records.")
                submitted = st.form_submit_button("Create User", type="primary", use_container_width=True)
                if submitted:
                    try:
                        service.add_user(
                            name=name,
                            role=role,
                            mobile=mobile,
                            password=password,
                            designation=designation,
                            details=details,
                            created_by=auth["name"],
                        )
                        st.success("User saved.")
                        st.rerun()
                    except ValueError as exc:
                        st.error(str(exc))

        with list_col:
            users = service.list_users()
            if users.empty:
                st.info("No users created yet.")
            else:
                display_users = users.rename(
                    columns={
                        "id": "User ID",
                        "name": "Name",
                        "mobile": "Mobile",
                        "role": "Role",
                        "designation": "Designation",
                        "details": "Details",
                        "active": "Active",
                        "has_password": "Observer Password",
                        "created_at": "Created At",
                        "created_by": "Created By",
                        "updated_at": "Updated At",
                    }
                )
                display_users["Role"] = display_users["Role"].map(_status_title)
                display_users["Observer Password"] = display_users.apply(
                    lambda row: "Not needed"
                    if row["Role"] == "Outsource"
                    else ("Yes" if row["Observer Password"] else "No"),
                    axis=1,
                )
                st.dataframe(display_users, use_container_width=True, hide_index=True)

                user_options = {
                    f"#{row['id']} | {row['name']} | {_status_title(row['role'])} | "
                    f"{'Active' if row['active'] else 'Inactive'}": int(row["id"])
                    for _, row in users.iterrows()
                }
                selected_user = st.selectbox(
                    "Activate or deactivate user",
                    options=list(user_options.keys()),
                    key="admin_toggle_user",
                )
                selected_row = users[users["id"] == user_options[selected_user]].iloc[0]
                next_state = not bool(selected_row["active"])
                action_col, edit_col = st.columns(2)
                with action_col:
                    if st.button(
                        "Reactivate User" if next_state else "Deactivate User",
                        use_container_width=True,
                        key="admin_toggle_user_btn",
                    ):
                        service.set_user_active(
                            user_id=int(selected_row["id"]),
                            active=next_state,
                            actor_name=auth["name"],
                        )
                        st.success("User status updated.")
                        st.rerun()

                with edit_col:
                    st.caption("Use the form below to reset password or correct details.")

                with st.expander("Edit Selected User / Reset Password", expanded=False):
                    with st.form("edit_attendance_user"):
                        edit_name = st.text_input("Name", value=str(selected_row["name"]))
                        edit_mobile = st.text_input("Mobile number", value=str(selected_row.get("mobile") or ""))
                        edit_role = st.selectbox(
                            "User type",
                            options=["observer", "outsource"],
                            index=0 if selected_row["role"] == "observer" else 1,
                            format_func=_status_title,
                        )
                        edit_designation = st.text_input(
                            "Designation / Work detail",
                            value=str(selected_row.get("designation") or ""),
                        )
                        if selected_row["role"] == "observer":
                            edit_password = st.text_input(
                                "New observer password",
                                type="password",
                                help="Leave blank to keep the current password.",
                            )
                        else:
                            edit_password = ""
                            st.caption("Outsource users do not need a password; mobile number is used.")
                        edit_details = st.text_area("Other details", value=str(selected_row.get("details") or ""))
                        update_submitted = st.form_submit_button(
                            "Update User",
                            type="primary",
                            use_container_width=True,
                        )
                        if update_submitted:
                            try:
                                service.update_user_profile(
                                    user_id=int(selected_row["id"]),
                                    name=edit_name,
                                    mobile=edit_mobile,
                                    role=edit_role,
                                    designation=edit_designation,
                                    details=edit_details,
                                    password=edit_password,
                                    actor_name=auth["name"],
                                )
                                st.success("User updated.")
                                st.rerun()
                            except ValueError as exc:
                                st.error(str(exc))

    with tabs[2]:
        st.subheader("Attendance Register")
        month = st.selectbox(
            "Attendance month",
            options=_month_options(service),
            format_func=_month_label,
            key="admin_att_month",
        )
        matrix = service.build_monthly_attendance_df(month)
        st.caption(
            "Legend: M 07:00-08:59, G 09:00-12:59, E 13:00-16:59, "
            "N 19:00-21:59, O other, P pending. Rejected entries are excluded."
        )
        st.dataframe(matrix, use_container_width=True, hide_index=True)

        st.markdown("---")
        st.subheader("Raw Login Data")
        st.dataframe(service.build_raw_attendance_df(month), use_container_width=True, hide_index=True)

    with tabs[3]:
        st.subheader("Excel Export")
        month = st.selectbox(
            "Export month",
            options=_month_options(service),
            format_func=_month_label,
            key="admin_export_month",
        )
        excel_bytes = service.export_attendance_workbook(month)
        st.download_button(
            "Download Attendance Excel",
            data=excel_bytes,
            file_name=f"outsource_attendance_{month}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    with tabs[4]:
        st.subheader("Audit Trail")
        st.dataframe(service.get_audit_log(limit=1000), use_container_width=True, hide_index=True)


def render_attendance_observer_page() -> None:
    service = get_attendance_service()
    auth = _require_user_auth(service, "observer")
    if not auth:
        return

    st.title("Observer Approval Desk")
    _render_metrics(service.get_summary_metrics())
    st.markdown("---")

    st.caption(f"Observer: {auth['name']}")
    month = st.selectbox(
        "Month",
        options=_month_options(service),
        format_func=_month_label,
        key="observer_month",
    )

    tabs = st.tabs(["Pending", "All Entries"])
    with tabs[0]:
        pending_entries = service.list_entries(month=month, status_filter="pending")
        st.dataframe(
            _display_entries(pending_entries) if not pending_entries.empty else _empty_entries_display(),
            use_container_width=True,
            hide_index=True,
        )
        actionable_entries = _observer_actionable_entries(pending_entries)
        _render_decision_form(
            service=service,
            entries=actionable_entries,
            actor_role="observer",
            actor_name=auth["name"],
            key_prefix="observer_pending",
        )

    with tabs[1]:
        entries = service.list_entries(month=month)
        st.dataframe(
            _display_entries(entries) if not entries.empty else _empty_entries_display(),
            use_container_width=True,
            hide_index=True,
        )
        st.markdown("---")
        st.caption("Completed entries are read-only for observers. Only pending entries can be decided.")
        editable_entries = _observer_actionable_entries(entries)
        _render_decision_form(
            service=service,
            entries=editable_entries,
            actor_role="observer",
            actor_name=auth["name"],
            key_prefix="observer_all",
        )


def render_outsource_login_page() -> None:
    service = get_attendance_service()
    auth = _require_user_auth(service, "outsource")
    if not auth:
        return

    current_time = now_ist()
    shift_code, shift_name = classify_shift(current_time)

    st.title("Outsource Login")
    st.caption(f"Outsource user: {auth['name']}")
    st.metric("Current IST Shift", f"{shift_code} - {shift_name}")
    st.caption(current_time.strftime("%d-%m-%Y %H:%M:%S IST"))

    with st.form("outsource_login_form"):
        st.text_input("Name", value=auth["name"], disabled=True)
        pc_name = st.text_input("PC Name", placeholder="Example: CYBER-PC-01")
        submitted = st.form_submit_button("Submit Login", type="primary", use_container_width=True)
        if submitted:
            try:
                entry_id = service.submit_login(int(auth["id"]), pc_name)
                st.success(f"Login submitted. Entry ID: {entry_id}")
            except ValueError as exc:
                st.error(str(exc))

    st.markdown("---")
    st.subheader("My Recent Entries")
    recent = service.list_entries(outsource_user_id=int(auth["id"])).head(10)
    st.dataframe(
        _display_entries(recent) if not recent.empty else _empty_entries_display(),
        use_container_width=True,
        hide_index=True,
    )
