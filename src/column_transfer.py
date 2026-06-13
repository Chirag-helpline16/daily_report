import io
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


SETTINGS_PATH = Path.home() / ".kiro" / "column_transfer_settings.json"
MATCH_MODES = {
    "Smart ID match": "smart",
    "Exact cleaned text": "exact",
}


def load_settings(path: Path = SETTINGS_PATH) -> Dict:
    try:
        if path.exists():
            return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return {}


def save_settings(settings: Dict, path: Path = SETTINGS_PATH) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(settings, indent=2, ensure_ascii=False), encoding="utf-8")


def _clean_text(value) -> str:
    if value is None:
        return ""
    try:
        if value != value:
            return ""
    except Exception:
        pass
    text = str(value).strip()
    if text.lower() in {"nan", "none", "<na>", "nat"}:
        return ""
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    if re.fullmatch(r"\d+(?:\.0+)?[eE]\+?\d+", text):
        try:
            return str(int(float(text)))
        except (OverflowError, ValueError):
            return text
    return re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]", "", text)


def normalize_match_value(value, mode: str = "smart") -> str:
    text = _clean_text(value)
    if mode == "exact":
        return re.sub(r"\s+", " ", text).upper()
    if re.fullmatch(r"[0-9\s\-_/]+\.0+", text):
        text = text.rsplit(".", 1)[0]
    return re.sub(r"[^A-Za-z0-9]+", "", text).upper()


def _unique_column_name(name: str, existing: set[str]) -> str:
    base_name = _clean_text(name) or "Added Column"
    candidate = base_name
    suffix = 2
    while candidate in existing:
        candidate = f"{base_name}_{suffix}"
        suffix += 1
    existing.add(candidate)
    return candidate


def transfer_columns(
    base_df: pd.DataFrame,
    lookup_df: pd.DataFrame,
    base_match_col: str,
    lookup_match_col: str,
    selected_lookup_cols: List[str],
    rename_map: Dict[str, str] | None = None,
    base_rename_map: Dict[str, str] | None = None,
    insert_after: str = "__END__",
    match_mode: str = "smart",
) -> Tuple[pd.DataFrame, Dict[str, int], Dict[str, str]]:
    if base_match_col not in base_df.columns:
        raise ValueError(f"File 1 match column not found: {base_match_col}")
    if lookup_match_col not in lookup_df.columns:
        raise ValueError(f"File 2 match column not found: {lookup_match_col}")
    missing = [column for column in selected_lookup_cols if column not in lookup_df.columns]
    if missing:
        raise ValueError(f"File 2 add column(s) not found: {', '.join(missing)}")

    rename_map = rename_map or {}
    base_rename_map = base_rename_map or {}
    base = base_df.copy()
    lookup = lookup_df[[lookup_match_col, *selected_lookup_cols]].copy()
    lookup_internal_cols = {
        source_column: f"_transfer_add_{idx}"
        for idx, source_column in enumerate(selected_lookup_cols)
    }
    lookup = lookup.rename(columns=lookup_internal_cols)

    base["_transfer_key"] = base[base_match_col].map(lambda value: normalize_match_value(value, match_mode))
    lookup["_transfer_key"] = lookup[lookup_match_col].map(lambda value: normalize_match_value(value, match_mode))
    base_blank_keys = int((base["_transfer_key"] == "").sum())
    lookup_blank_keys = int((lookup["_transfer_key"] == "").sum())

    lookup = lookup[lookup["_transfer_key"] != ""].copy()
    before_lookup_dedupe = len(lookup)
    lookup = lookup.drop_duplicates(subset="_transfer_key", keep="first")
    duplicate_lookup_keys = before_lookup_dedupe - len(lookup)

    merged = base.merge(
        lookup[["_transfer_key", *lookup_internal_cols.values()]],
        on="_transfer_key",
        how="left",
        suffixes=("", "_from_file2"),
    )

    existing_names = set(base_df.columns)
    final_add_names: Dict[str, str] = {}
    added_data = {}
    for source_column in selected_lookup_cols:
        output_name = _unique_column_name(rename_map.get(source_column, source_column), existing_names)
        final_add_names[source_column] = output_name
        added_data[output_name] = merged[lookup_internal_cols[source_column]].map(_clean_text)

    result = base_df.copy()
    insert_index = len(result.columns)
    if insert_after == "__START__":
        insert_index = 0
    elif insert_after != "__END__" and insert_after in result.columns:
        insert_index = list(result.columns).index(insert_after) + 1

    for offset, output_name in enumerate(final_add_names.values()):
        result.insert(insert_index + offset, output_name, added_data[output_name])

    final_columns: List[str] = []
    used_final_names: set[str] = set()
    final_added_names: Dict[str, str] = {}
    preliminary_added_sources = {name: source for source, name in final_add_names.items()}

    for column in result.columns:
        if column in base_df.columns:
            desired_name = base_rename_map.get(column, column)
        else:
            desired_name = column
        final_name = _unique_column_name(desired_name, used_final_names)
        final_columns.append(final_name)
        if column in preliminary_added_sources:
            final_added_names[preliminary_added_sources[column]] = final_name

    result.columns = final_columns
    final_add_names = final_added_names

    matched_rows = int(merged["_transfer_key"].isin(set(lookup["_transfer_key"])).sum())
    stats = {
        "file1_rows": len(base_df),
        "file2_rows": len(lookup_df),
        "matched_rows": matched_rows,
        "unmatched_rows": len(base_df) - matched_rows,
        "base_blank_keys": base_blank_keys,
        "lookup_blank_keys": lookup_blank_keys,
        "duplicate_lookup_keys": duplicate_lookup_keys,
        "columns_added": len(selected_lookup_cols),
        "base_columns_renamed": sum(
            1
            for column, new_name in base_rename_map.items()
            if column in base_df.columns and _clean_text(new_name) and _clean_text(new_name) != column
        ),
    }
    return result, stats, final_add_names


def _read_uploaded_file(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name.lower()
    uploaded_file.seek(0)
    if name.endswith(".csv"):
        return pd.read_csv(uploaded_file, dtype=str, keep_default_na=False)
    return pd.read_excel(uploaded_file, dtype=str, keep_default_na=False)


def _styled_excel_bytes(df: pd.DataFrame) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data")
        worksheet = writer.sheets["Data"]

        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        body_font = Font(name="Calibri", size=10, color="111827")
        border_side = Side(style="thin", color="D9E2F3")
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="top", wrap_text=True)
        even_fill = PatternFill("solid", fgColor="F4F8FB")
        odd_fill = PatternFill("solid", fgColor="FFFFFF")

        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = False
        worksheet.row_dimensions[1].height = 28

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            row_fill = even_fill if row_idx % 2 == 0 else odd_fill
            for cell in row:
                cell.font = body_font
                cell.border = border
                cell.fill = row_fill
                cell.alignment = left

        for column_cells in worksheet.columns:
            letter = column_cells[0].column_letter
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[letter].width = min(max(max_length + 3, 12), 48)

        if worksheet.max_row > 1 and worksheet.max_column > 1:
            table_ref = f"A1:{worksheet.cell(row=1, column=worksheet.max_column).column_letter}{worksheet.max_row}"
            worksheet.auto_filter.ref = table_ref
            table = Table(displayName="ColumnTransferTable", ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=False,
                showColumnStripes=False,
            )
            worksheet.add_table(table)

    return buffer.getvalue()


def _default_index(options: List[str], saved_value: str | None, fallback: str | None = None) -> int:
    if saved_value in options:
        return options.index(saved_value)
    if fallback in options:
        return options.index(fallback)
    return 0


def _guess_match_column(columns: List[str]) -> str | None:
    aliases = [
        "Acknowledgement No.",
        "Acknowledgement Number",
        "Ack No",
        "ACK No",
        "Account No.",
        "Account Number",
        "Bank Account No",
    ]
    normalized = {re.sub(r"[^a-z0-9]+", "", column.lower()): column for column in columns}
    for alias in aliases:
        key = re.sub(r"[^a-z0-9]+", "", alias.lower())
        if key in normalized:
            return normalized[key]
    return columns[0] if columns else None


def render_column_transfer_page():
    st.title("Add Columns From Another File")

    settings = load_settings()
    with st.expander("Saved Settings", expanded=False):
        st.caption(str(SETTINGS_PATH))
        if st.button("Clear saved settings", key="column_transfer_clear"):
            save_settings({})
            st.success("Saved settings cleared.")
            st.rerun()

    col1, col2 = st.columns(2)
    with col1:
        base_file = st.file_uploader(
            "File 1 - Base file",
            type=["xlsx", "xls", "csv"],
            key="column_transfer_base_file",
        )
    with col2:
        lookup_file = st.file_uploader(
            "File 2 - Add columns from this file",
            type=["xlsx", "xls", "csv"],
            key="column_transfer_lookup_file",
        )

    if base_file is None or lookup_file is None:
        st.info("Upload both files to start matching and adding columns.")
        return

    try:
        base_df = _read_uploaded_file(base_file)
        lookup_df = _read_uploaded_file(lookup_file)

        st.success(f"File 1 loaded: {len(base_df):,} rows | File 2 loaded: {len(lookup_df):,} rows")

        base_columns = list(base_df.columns)
        lookup_columns = list(lookup_df.columns)
        match_labels = list(MATCH_MODES.keys())
        saved_mode_label = next(
            (label for label, value in MATCH_MODES.items() if value == settings.get("match_mode")),
            "Smart ID match",
        )

        st.markdown("### Match Settings")
        match_col1, match_col2, match_col3 = st.columns(3)
        with match_col1:
            base_match_col = st.selectbox(
                "File 1 match column",
                options=base_columns,
                index=_default_index(base_columns, settings.get("base_match_col"), _guess_match_column(base_columns)),
                key="column_transfer_base_match",
            )
        with match_col2:
            lookup_match_col = st.selectbox(
                "File 2 match column",
                options=lookup_columns,
                index=_default_index(lookup_columns, settings.get("lookup_match_col"), _guess_match_column(lookup_columns)),
                key="column_transfer_lookup_match",
            )
        with match_col3:
            match_mode_label = st.selectbox(
                "Match mode",
                options=match_labels,
                index=match_labels.index(saved_mode_label),
                key="column_transfer_match_mode",
            )
        match_mode = MATCH_MODES[match_mode_label]

        st.markdown("### Columns To Add")
        add_candidates = [column for column in lookup_columns if column != lookup_match_col]
        saved_selected = [column for column in settings.get("selected_lookup_cols", []) if column in add_candidates]
        selected_lookup_cols = st.multiselect(
            "Select File 2 columns to add into File 1",
            options=add_candidates,
            default=saved_selected,
            key="column_transfer_selected_columns",
        )

        if not selected_lookup_cols:
            st.warning("Select at least one File 2 column to add.")
            return

        insert_options = ["At beginning", "At end"] + base_columns
        saved_insert = settings.get("insert_after", "__END__")
        saved_insert_label = (
            "At beginning"
            if saved_insert == "__START__"
            else "At end"
            if saved_insert == "__END__"
            else saved_insert
        )
        insert_label = st.selectbox(
            "Insert added columns after",
            options=insert_options,
            index=_default_index(insert_options, saved_insert_label, "At end"),
            key="column_transfer_insert_after",
        )
        insert_after = "__START__" if insert_label == "At beginning" else "__END__" if insert_label == "At end" else insert_label

        st.markdown("### Rename Added Columns")
        saved_renames = settings.get("rename_map", {})
        rename_map: Dict[str, str] = {}
        rename_cols = st.columns(2)
        for idx, source_column in enumerate(selected_lookup_cols):
            with rename_cols[idx % 2]:
                rename_map[source_column] = st.text_input(
                    f"{source_column}",
                    value=saved_renames.get(source_column, source_column),
                    key=f"column_transfer_rename_{idx}_{re.sub(r'[^A-Za-z0-9_]+', '_', source_column)}",
                )

        base_rename_map: Dict[str, str] = {}
        saved_base_renames = settings.get("base_rename_map", {})
        with st.expander("Rename File 1 Columns", expanded=False):
            st.caption("These names apply to the final output file. Matching and insert position still use the original File 1 headers.")
            base_rename_cols = st.columns(2)
            for idx, base_column in enumerate(base_columns):
                with base_rename_cols[idx % 2]:
                    base_rename_map[base_column] = st.text_input(
                        f"{base_column}",
                        value=saved_base_renames.get(base_column, base_column),
                        key=f"column_transfer_base_rename_{idx}_{re.sub(r'[^A-Za-z0-9_]+', '_', base_column)}",
                    )

        current_settings = {
            "base_match_col": base_match_col,
            "lookup_match_col": lookup_match_col,
            "match_mode": match_mode,
            "selected_lookup_cols": selected_lookup_cols,
            "rename_map": rename_map,
            "base_rename_map": base_rename_map,
            "insert_after": insert_after,
        }
        save_settings(current_settings)

        output_df, stats, final_names = transfer_columns(
            base_df=base_df,
            lookup_df=lookup_df,
            base_match_col=base_match_col,
            lookup_match_col=lookup_match_col,
            selected_lookup_cols=selected_lookup_cols,
            rename_map=rename_map,
            base_rename_map=base_rename_map,
            insert_after=insert_after,
            match_mode=match_mode,
        )

        st.markdown("### Output Summary")
        metric_cols = st.columns(5)
        metric_cols[0].metric("File 1 Rows", f"{stats['file1_rows']:,}")
        metric_cols[1].metric("Matched Rows", f"{stats['matched_rows']:,}")
        metric_cols[2].metric("Unmatched Rows", f"{stats['unmatched_rows']:,}")
        metric_cols[3].metric("Columns Added", f"{stats['columns_added']:,}")
        metric_cols[4].metric("Duplicate File 2 Keys", f"{stats['duplicate_lookup_keys']:,}")

        if stats["base_columns_renamed"]:
            st.caption(f"File 1 columns renamed in output: {stats['base_columns_renamed']:,}")

        if any(final_names[column] != (rename_map.get(column) or column) for column in selected_lookup_cols):
            st.warning("Some added column names were adjusted to avoid duplicate output headers.")

        st.dataframe(output_df.head(100), use_container_width=True, hide_index=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_bytes = _styled_excel_bytes(output_df)
        csv_bytes = output_df.to_csv(index=False).encode("utf-8-sig")

        dl1, dl2 = st.columns(2)
        with dl1:
            st.download_button(
                "Download Excel",
                data=excel_bytes,
                file_name=f"columns_added_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
                key="column_transfer_download_excel",
            )
        with dl2:
            st.download_button(
                "Download CSV",
                data=csv_bytes,
                file_name=f"columns_added_{timestamp}.csv",
                mime="text/csv",
                use_container_width=True,
                key="column_transfer_download_csv",
            )
    except Exception as exc:
        st.error(f"Error processing files: {exc}")
