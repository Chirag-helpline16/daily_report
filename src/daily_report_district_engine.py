"""Daily report district split engine.

Ported from the standalone IFSC District Splitter. This module keeps the
processing logic available to Streamlit without using the standalone app's
uploads/downloads folders.
"""
from __future__ import annotations

import io
import os
import re
import sqlite3
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from functools import lru_cache
from pathlib import Path
from typing import Any
from urllib.parse import quote

import pandas as pd
import requests
from requests.exceptions import SSLError
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    import config
except ImportError:  # pragma: no cover - keeps app.py portable
    config = None


BASE_DIR = Path(__file__).resolve().parents[1]


def _setting(name: str, default: Any) -> Any:
    if config and hasattr(config, name):
        return getattr(config, name)
    return default


def _path_setting(name: str, default: str) -> Path:
    value = Path(os.getenv(name, str(_setting(name, default))))
    return value if value.is_absolute() else BASE_DIR / value


UPLOAD_FOLDER = _path_setting("UPLOAD_FOLDER", "uploads")
DOWNLOAD_FOLDER = _path_setting("DOWNLOAD_FOLDER", "downloads")
ALLOWED_EXTENSIONS = set(_setting("ALLOWED_EXTENSIONS", {"xlsx", "xls", "csv"}))
MAX_FILE_SIZE = int(os.getenv("MAX_FILE_SIZE", _setting("MAX_FILE_SIZE", 50 * 1024 * 1024)))

IFSC_API_BASE = os.getenv("IFSC_API_BASE", _setting("IFSC_API_BASE", "https://ifsc.razorpay.com")).rstrip("/")
IFSC_API_TIMEOUT = int(os.getenv("IFSC_API_TIMEOUT", _setting("IFSC_API_TIMEOUT", 10)))

# Set POSTAL_API_BASE_URL to a deployed nstack-in/indian-postal-code-api host if you have one.
POSTAL_API_BASE_URL = os.getenv("POSTAL_API_BASE_URL", _setting("POSTAL_API_BASE_URL", "")).rstrip("/")
PUBLIC_PINCODE_URL_TEMPLATE = os.getenv(
    "PUBLIC_PINCODE_URL_TEMPLATE",
    _setting("PUBLIC_PINCODE_URL_TEMPLATE", "https://api.postalpincode.in/pincode/{pincode}"),
)
POSTAL_API_TIMEOUT = int(os.getenv("POSTAL_API_TIMEOUT", _setting("POSTAL_API_TIMEOUT", 10)))
POSTAL_DB_URL = os.getenv(
    "POSTAL_DB_URL",
    _setting(
        "POSTAL_DB_URL",
        "https://raw.githubusercontent.com/nstack-in/indian-postal-code-api/master/db/data.sqlite",
    ),
)
POSTAL_DB_PATH = _path_setting("POSTAL_DB_PATH", "data/indian_postal_code.sqlite")
SSL_VERIFY = str(os.getenv("SSL_VERIFY", _setting("SSL_VERIFY", "true"))).lower() in {"1", "true", "yes", "on"}
SSL_FALLBACK_VERIFY_FALSE = str(
    os.getenv("SSL_FALLBACK_VERIFY_FALSE", _setting("SSL_FALLBACK_VERIFY_FALSE", "true"))
).lower() in {"1", "true", "yes", "on"}

FLASK_HOST = os.getenv("FLASK_HOST", _setting("FLASK_HOST", "0.0.0.0"))
FLASK_PORT = int(os.getenv("PORT", os.getenv("FLASK_PORT", _setting("FLASK_PORT", 5001))))
FLASK_DEBUG = str(os.getenv("FLASK_DEBUG", _setting("FLASK_DEBUG", "true"))).lower() in {"1", "true", "yes", "on"}

GUJARAT_STATE_CODE = "IN-GJ"
GUJARAT_STATE_NAMES = {"GUJARAT", "GUJRAT", "GJ", "IN GJ", "IN-GJ"}

CANONICAL_GUJARAT_DISTRICTS = [
    "Ahmedabad",
    "Amreli",
    "Anand",
    "Aravalli",
    "Banaskantha",
    "Bharuch",
    "Bhavnagar",
    "Botad",
    "Chhota Udaipur",
    "Dahod",
    "Dang",
    "Devbhumi Dwarka",
    "Gandhinagar",
    "Gir Somnath",
    "Jamnagar",
    "Junagadh",
    "Kachchh",
    "Kheda",
    "Mahisagar",
    "Mehsana",
    "Morbi",
    "Narmada",
    "Navsari",
    "Panchmahal",
    "Patan",
    "Porbandar",
    "Rajkot",
    "Sabarkantha",
    "Surat",
    "Surendranagar",
    "Tapi",
    "Vadodara",
    "Valsad",
    "Vav-Tharad",
]


def normalize_text(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = str(value).strip().upper()
    text = text.replace("&", " AND ")
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


DISTRICT_ALIASES = {
    normalize_text(name): name for name in CANONICAL_GUJARAT_DISTRICTS
}
DISTRICT_ALIASES.update(
    {
        "AHMADABAD": "Ahmedabad",
        "AHMEDABAD CITY": "Ahmedabad",
        "AHMEDABAD RURAL": "Ahmedabad",
        "BANAS KANTHA": "Banaskantha",
        "BANAS KANTHA DISTRICT": "Banaskantha",
        "BANA KANTHA": "Banaskantha",
        "BARODA": "Vadodara",
        "CHHOTA UDEPUR": "Chhota Udaipur",
        "CHHOTAUDEPUR": "Chhota Udaipur",
        "CHHOTA UDAIPUR": "Chhota Udaipur",
        "DANGS": "Dang",
        "THE DANGS": "Dang",
        "DEVBHOOMI DWARKA": "Devbhumi Dwarka",
        "DEVBHUMI DWARKA": "Devbhumi Dwarka",
        "DWARKA": "Devbhumi Dwarka",
        "GIR SOMANATH": "Gir Somnath",
        "GIRSOMNATH": "Gir Somnath",
        "JUNAGARH": "Junagadh",
        "KACHH": "Kachchh",
        "KACHCH": "Kachchh",
        "KUCHCH": "Kachchh",
        "KUCHCHH": "Kachchh",
        "KUTCH": "Kachchh",
        "KUTCHH": "Kachchh",
        "MAHI SAGAR": "Mahisagar",
        "MAHISAGER": "Mahisagar",
        "MAHESANA": "Mehsana",
        "MEHASANA": "Mehsana",
        "PANCH MAHAL": "Panchmahal",
        "PANCH MAHALS": "Panchmahal",
        "PANCHMAHALS": "Panchmahal",
        "SABAR KANTHA": "Sabarkantha",
        "SABARKANTHA DISTRICT": "Sabarkantha",
        "SURAT CITY": "Surat",
        "SURAT RURAL": "Surat",
        "SURENDRA NAGAR": "Surendranagar",
        "VADODRA": "Vadodara",
        "VADODARA RURAL": "Vadodara",
        "VAV THARAD": "Vav-Tharad",
        "VAVTHARAD": "Vav-Tharad",
    }
)

LOCATION_DISTRICT_ALIASES = {
    "BAREJA": "Ahmedabad",
    "BHUJ": "Kachchh",
    "DASKOI": "Ahmedabad",
    "DHARI": "Amreli",
    "GONDAL": "Rajkot",
    "MORVI": "Morbi",
    "MUNDRA": "Kachchh",
    "NADIAD": "Kheda",
    "SALUN": "Kheda",
    "SUTRAPADA": "Gir Somnath",
    "TANKARA": "Morbi",
    "THARAD": "Vav-Tharad",
    "UMRETH": "Anand",
}

IFSC_COLUMN_CANDIDATES = {"IFSC", "IFSCCODE", "IFSCCODENO", "IFSCCODENUMBER"}
PINCODE_COLUMN_CANDIDATES = {"PIN", "PINCODE", "POSTALCODE", "POSTCODE", "ZIP", "ZIPCODE"}
IFSC_RE = re.compile(r"^[A-Z]{4}0[A-Z0-9]{6}$")
PINCODE_RE = re.compile(r"\b\d{3}\s?\d{3}\b")
UPLOAD_ID_RE = re.compile(r"^[a-f0-9]{32}$")
OTHER_GROUP_NAME = "Other"
OUTPUT_COLUMN_SOURCES = [
    ("S No.", ("S No.", "S No", "Sr No.", "Sr No")),
    ("Acknowledgement No.", ("Acknowledgement No.", "Acknowledgement No", "Ack No.", "Ack No")),
    ("Suspect District", ("Suspect District", "District")),
    ("Suspect Account No.", ("Suspect Account No.", "Account No.", "Account No")),
    ("IFSC Code", ("IFSC Code", "IFSC")),
    ("Address", ("Address",)),
    ("Pin Code", ("Pin Code", "Pincode", "PIN", "Pin")),
    ("Transaction Amount", ("Transaction Amount",)),
    ("Disputed Amount", ("Disputed Amount",)),
    ("Bank/FIs", ("Bank/FIs", "Bank/FI", "Bank FIs")),
    ("Layers", ("Layers", "Layer")),
    ("Victim District", ("Victim District",)),
    ("Reported Amount (Victim)", ("Reported Amount (Victim)",)),
]
OUTPUT_COLUMNS = [column for column, _sources in OUTPUT_COLUMN_SOURCES]

def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def canonicalize_district(value: Any) -> str | None:
    return DISTRICT_ALIASES.get(normalize_text(value))


def mapped_location_district(value: Any) -> str | None:
    return LOCATION_DISTRICT_ALIASES.get(normalize_text(value))


def district_from_value(value: Any, allow_location: bool = False) -> str | None:
    return canonicalize_district(value) or (mapped_location_district(value) if allow_location else None)


def district_from_explicit_text(value: Any) -> str | None:
    normalized = normalize_text(value)
    if not normalized:
        return None

    for alias, district in sorted(DISTRICT_ALIASES.items(), key=lambda item: len(item[0]), reverse=True):
        if re.search(rf"\b(?:DIST|DISTRICT)\s+{re.escape(alias)}\b", normalized):
            return district
    return None


def extract_pincodes_from_text(*values: Any) -> list[str]:
    pincodes: list[str] = []
    for value in values:
        for match in PINCODE_RE.findall(str(value or "")):
            pincode = clean_pincode(match)
            if pincode and pincode not in pincodes:
                pincodes.append(pincode)
    return pincodes


def first_row_value(row: dict[str, Any], candidates: tuple[str | None, ...]) -> Any:
    for candidate in candidates:
        if candidate and candidate in row:
            return row.get(candidate, "")
    return ""


def build_output_row(original: dict[str, Any], ifsc_column: str | None, pincode_column: str | None) -> dict[str, Any]:
    output_row: dict[str, Any] = {}
    for output_column, source_columns in OUTPUT_COLUMN_SOURCES:
        candidates: tuple[str | None, ...] = source_columns
        if output_column == "IFSC Code":
            candidates = (ifsc_column, *source_columns)
        elif output_column == "Pin Code":
            candidates = (pincode_column, *source_columns)
        output_row[output_column] = first_row_value(original, candidates)
    return output_row


def safe_get(url: str, timeout: int) -> requests.Response:
    try:
        return requests.get(url, timeout=timeout, headers={"Accept": "application/json"}, verify=SSL_VERIFY)
    except SSLError:
        if not SSL_FALLBACK_VERIFY_FALSE:
            raise
        try:
            requests.packages.urllib3.disable_warnings()  # type: ignore[attr-defined]
        except Exception:
            pass
        return requests.get(url, timeout=timeout, headers={"Accept": "application/json"}, verify=False)


def safe_get_binary(url: str, timeout: int) -> requests.Response:
    try:
        return requests.get(url, timeout=timeout, verify=SSL_VERIFY)
    except SSLError:
        if not SSL_FALLBACK_VERIFY_FALSE:
            raise
        try:
            requests.packages.urllib3.disable_warnings()  # type: ignore[attr-defined]
        except Exception:
            pass
        return requests.get(url, timeout=timeout, verify=False)


def is_gujarat_state(state: Any = None, iso3166: Any = None) -> bool:
    iso = str(iso3166 or "").strip().upper()
    if iso == GUJARAT_STATE_CODE:
        return True
    return normalize_text(state) in GUJARAT_STATE_NAMES


def clean_ifsc(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return re.sub(r"\s+", "", str(value).strip().upper())


def clean_pincode(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    digits = re.sub(r"\D", "", text)
    return digits if len(digits) == 6 else ""


def find_column(columns: list[str], candidates: set[str]) -> str | None:
    for column in columns:
        normalized = re.sub(r"[^A-Z0-9]", "", str(column).upper())
        if normalized in candidates:
            return column
    for column in columns:
        normalized = re.sub(r"[^A-Z0-9]", "", str(column).upper())
        if "IFSC" in normalized and candidates is IFSC_COLUMN_CANDIDATES:
            return column
        if ("PIN" in normalized or "POSTAL" in normalized) and candidates is PINCODE_COLUMN_CANDIDATES:
            return column
    return None


def unique_column_name(existing: set[str], preferred: str) -> str:
    if preferred not in existing:
        existing.add(preferred)
        return preferred
    index = 2
    while f"{preferred}_{index}" in existing:
        index += 1
    name = f"{preferred}_{index}"
    existing.add(name)
    return name


def read_input_file(filepath: Path) -> pd.DataFrame:
    suffix = filepath.suffix.lower()
    if suffix == ".csv":
        df = pd.read_csv(filepath, dtype=str, keep_default_na=False)
    elif suffix in {".xlsx", ".xls"}:
        df = pd.read_excel(filepath, dtype=str, keep_default_na=False)
    else:
        raise ValueError("Unsupported file type")

    if df.empty:
        raise ValueError("The uploaded file has no rows")

    df.columns = [str(column).strip() for column in df.columns]
    return df


@lru_cache(maxsize=10000)
def lookup_ifsc(ifsc_code: str) -> dict[str, Any]:
    if not ifsc_code:
        return {"ok": False, "reason": "IFSC missing"}
    if not IFSC_RE.match(ifsc_code):
        return {"ok": False, "reason": "IFSC format is invalid"}

    url = f"{IFSC_API_BASE}/{quote(ifsc_code)}"
    try:
        response = safe_get(url, IFSC_API_TIMEOUT)
    except requests.RequestException as exc:
        return {"ok": False, "reason": f"IFSC API error: {exc.__class__.__name__}"}

    if response.status_code == 404:
        return {"ok": False, "reason": "IFSC not found"}
    if response.status_code != 200:
        return {"ok": False, "reason": f"IFSC API returned HTTP {response.status_code}"}

    try:
        payload = response.json()
    except ValueError:
        return {"ok": False, "reason": "IFSC API returned invalid JSON"}

    raw_district = payload.get("DISTRICT")
    city = payload.get("CITY")
    branch = payload.get("BRANCH")
    address = payload.get("ADDRESS")
    district = (
        district_from_value(raw_district)
        or district_from_value(raw_district, allow_location=True)
        or district_from_explicit_text(branch)
        or district_from_explicit_text(address)
        or district_from_value(city)
        or district_from_value(city, allow_location=True)
    )
    state = payload.get("STATE")
    iso3166 = payload.get("ISO3166")

    if not district:
        for pincode in extract_pincodes_from_text(address, branch):
            pincode_result = lookup_pincode_local_db(pincode)
            if pincode_result.get("ok"):
                district = pincode_result["district"]
                break

    ok = bool(district and is_gujarat_state(state, iso3166))
    reason = None
    if not district:
        reason = f"IFSC district was not a recognized district: {raw_district or 'blank'}"
    elif not is_gujarat_state(state, iso3166):
        reason = f"IFSC state is not Gujarat: {state or iso3166 or 'blank'}"

    return {
        "ok": ok,
        "district": district,
        "raw_district": raw_district,
        "city": city,
        "state": state,
        "iso3166": iso3166,
        "bank": payload.get("BANK"),
        "branch": branch,
        "address": address,
        "reason": reason,
    }


def build_nstack_pincode_url(pincode: str) -> str | None:
    if not POSTAL_API_BASE_URL:
        return None
    base = POSTAL_API_BASE_URL.rstrip("/")
    if "{pincode}" in base:
        return base.format(pincode=quote(pincode))
    if base.endswith("/find"):
        return f"{base}/{quote(pincode)}"
    if base.endswith("/api/pin"):
        return f"{base}/find/{quote(pincode)}"
    return f"{base}/api/pin/find/{quote(pincode)}"


def iter_postal_objects(payload: Any) -> list[dict[str, Any]]:
    objects: list[dict[str, Any]] = []

    def walk(node: Any) -> None:
        if isinstance(node, list):
            for item in node:
                walk(item)
        elif isinstance(node, dict):
            objects.append(node)
            for key in ("data", "PostOffice", "postOffice", "post_offices", "value", "results"):
                if key in node:
                    walk(node[key])

    walk(payload)
    return objects


def extract_postal_candidates(payload: Any) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    for item in iter_postal_objects(payload):
        district = (
            item.get("District")
            or item.get("district")
            or item.get("DistrictsName")
            or item.get("districtsName")
            or item.get("DistrictName")
            or item.get("district_name")
        )
        state = item.get("State") or item.get("state") or item.get("Circle") or item.get("circle")
        if district:
            candidates.append({"district": district, "state": state})
    return candidates


def ensure_postal_db() -> None:
    if POSTAL_DB_PATH.exists() and POSTAL_DB_PATH.stat().st_size > 0:
        return

    POSTAL_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    response = safe_get_binary(POSTAL_DB_URL, POSTAL_API_TIMEOUT)
    if response.status_code != 200 or not response.content:
        raise RuntimeError(f"postal database download returned HTTP {response.status_code}")

    temp_path = POSTAL_DB_PATH.with_suffix(".tmp")
    temp_path.write_bytes(response.content)
    temp_path.replace(POSTAL_DB_PATH)


def lookup_pincode_local_db(pincode: str) -> dict[str, Any]:
    try:
        ensure_postal_db()
        with sqlite3.connect(POSTAL_DB_PATH) as connection:
            rows = connection.execute(
                """
                SELECT PostOfficeName, DistrictsName, City, State
                FROM pincodes
                WHERE Pincode = ?
                """,
                (pincode,),
            ).fetchall()
    except Exception as exc:
        return {"ok": False, "reason": f"postal database lookup failed: {exc}"}

    for post_office, raw_district, city, state in rows:
        district = mapped_location_district(post_office)
        if district and is_gujarat_state(state):
            return {
                "ok": True,
                "district": district,
                "raw_district": raw_district,
                "post_office": post_office,
                "city": city,
                "state": state,
                "source": "postal_database",
            }

    for post_office, raw_district, city, state in rows:
        district = mapped_location_district(city) or district_from_value(raw_district, allow_location=True)
        if district and is_gujarat_state(state):
            return {
                "ok": True,
                "district": district,
                "raw_district": raw_district,
                "post_office": post_office,
                "city": city,
                "state": state,
                "source": "postal_database",
            }

    return {"ok": False, "reason": "postal database did not return a recognized Gujarat district"}


def fetch_json(url: str, timeout: int) -> tuple[int, Any | None]:
    try:
        response = safe_get(url, timeout)
    except requests.RequestException:
        return 0, None
    try:
        payload = response.json()
    except ValueError:
        payload = None
    return response.status_code, payload


@lru_cache(maxsize=10000)
def lookup_pincode(pincode: str) -> dict[str, Any]:
    if not pincode:
        return {"ok": False, "reason": "Pincode missing"}

    sources: list[tuple[str, str]] = []
    nstack_url = build_nstack_pincode_url(pincode)
    if nstack_url:
        sources.append(("nstack_postal_api", nstack_url))
    if PUBLIC_PINCODE_URL_TEMPLATE:
        sources.append(("public_pincode_api", PUBLIC_PINCODE_URL_TEMPLATE.format(pincode=quote(pincode))))

    last_reason = "No pincode source configured"
    for source_name, url in sources:
        status_code, payload = fetch_json(url, POSTAL_API_TIMEOUT)
        if status_code != 200 or payload is None:
            last_reason = f"{source_name} returned HTTP {status_code or 'error'}"
            continue

        for candidate in extract_postal_candidates(payload):
            district = district_from_value(candidate.get("district"), allow_location=True)
            state = candidate.get("state")
            if district and (not state or is_gujarat_state(state)):
                return {
                    "ok": True,
                    "district": district,
                    "raw_district": candidate.get("district"),
                    "state": state,
                    "source": source_name,
                }
        last_reason = f"{source_name} did not return a recognized Gujarat district"

    local_result = lookup_pincode_local_db(pincode)
    if local_result.get("ok"):
        return local_result

    return {"ok": False, "reason": f"{last_reason}; {local_result.get('reason')}"}


def resolve_row(ifsc_code: str, pincode: str) -> dict[str, Any]:
    ifsc_result = lookup_ifsc(ifsc_code) if ifsc_code else {"ok": False, "reason": "IFSC missing"}
    if ifsc_result.get("ok"):
        return {
            "ok": True,
            "district": ifsc_result["district"],
            "source": "IFSC",
            "ifsc_result": ifsc_result,
            "pincode_result": None,
        }

    pincode_result = lookup_pincode(pincode) if pincode else {"ok": False, "reason": "Pincode missing"}
    if pincode_result.get("ok"):
        return {
            "ok": True,
            "district": pincode_result["district"],
            "source": "Pincode",
            "ifsc_result": ifsc_result,
            "pincode_result": pincode_result,
        }

    return {
        "ok": False,
        "district": None,
        "source": None,
        "ifsc_result": ifsc_result,
        "pincode_result": pincode_result,
        "reason": f"{ifsc_result.get('reason', 'IFSC failed')}; {pincode_result.get('reason', 'pincode failed')}",
    }


def resolve_mapped_column(columns: list[str], value: Any, candidates: set[str]) -> str | None:
    if value is None or str(value).strip() == "__auto__":
        return find_column(columns, candidates)
    text = str(value).strip()
    if not text or text == "__none__":
        return None
    if text not in columns:
        raise ValueError(f"Selected column was not found: {text}")
    return text


def process_input_file(
    filepath: Path,
    ifsc_column: str | None = None,
    pincode_column: str | None = None,
) -> tuple[dict[str, list[dict[str, Any]]], list[dict[str, Any]], dict[str, Any]]:
    df = read_input_file(filepath)
    columns = list(df.columns)
    ifsc_column = resolve_mapped_column(columns, ifsc_column, IFSC_COLUMN_CANDIDATES)
    pincode_column = resolve_mapped_column(columns, pincode_column, PINCODE_COLUMN_CANDIDATES)

    if not ifsc_column and not pincode_column:
        raise ValueError("Please map at least one column: IFSC or Pincode")

    district_data: dict[str, list[dict[str, Any]]] = defaultdict(list)
    other_rows: list[dict[str, Any]] = []
    unmatched_details: list[dict[str, Any]] = []
    source_counter: Counter[str] = Counter()

    for index, row in df.iterrows():
        row_number = int(index) + 2
        original = row.to_dict()
        output_row = build_output_row(original, ifsc_column, pincode_column)
        ifsc_code = clean_ifsc(original.get(ifsc_column)) if ifsc_column else ""
        pincode = clean_pincode(original.get(pincode_column)) if pincode_column else ""
        result = resolve_row(ifsc_code, pincode)

        if result["ok"]:
            district = result["district"]
            district_data[district].append(output_row)
            source_counter[result["source"]] += 1
            continue

        other_rows.append(output_row)
        unmatched_details.append(
            {
                "row": row_number,
                "ifsc": ifsc_code,
                "pincode": pincode,
                "reason": result.get("reason", "Could not determine district"),
            }
        )

    if other_rows:
        district_data[OTHER_GROUP_NAME] = other_rows

    matched_rows = int(sum(len(rows) for district, rows in district_data.items() if district != OTHER_GROUP_NAME))
    other_count = int(len(other_rows))
    summary = {
        "input_rows": int(len(df)),
        "matched_rows": matched_rows,
        "other_rows": other_count,
        "skipped_rows": other_count,
        "total_districts": int(len([district for district in district_data if district != OTHER_GROUP_NAME])),
        "total_output_files": int(len(district_data)),
        "matched_by_ifsc": int(source_counter["IFSC"]),
        "matched_by_pincode": int(source_counter["Pincode"]),
        "ifsc_column": ifsc_column,
        "pincode_column": pincode_column,
        "columns": OUTPUT_COLUMNS,
        "districts": {
            district: {"record_count": len(rows)}
            for district, rows in sorted(district_data.items())
        },
    }
    return dict(district_data), unmatched_details, summary


def safe_stem(value: str) -> str:
    stem = re.sub(r"[^A-Za-z0-9]+", " ", value).strip()
    stem = re.sub(r"\s+", " ", stem)
    return stem.upper() or "DISTRICT"


def ordered_group_items(grouped_data: dict[str, list[dict[str, Any]]]) -> list[tuple[str, list[dict[str, Any]]]]:
    return sorted(grouped_data.items(), key=lambda item: (item[0] == OTHER_GROUP_NAME, safe_stem(item[0])))


def style_output_worksheet(worksheet) -> None:
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    green_fill = PatternFill("solid", fgColor="D9EAD3")
    red_fill = PatternFill("solid", fgColor="F4CCCC")
    header_font = Font(bold=True, color="000000")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_gray = Side(style="thin", color="D9E2F3")
    header_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    worksheet.freeze_panes = "A2"
    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions

    for column_index in range(1, worksheet.max_column + 1):
        header_cell = worksheet.cell(row=1, column=column_index)
        if column_index <= 2:
            header_cell.fill = yellow_fill
        elif column_index > max(worksheet.max_column - 2, 2):
            header_cell.fill = green_fill
        else:
            header_cell.fill = red_fill
        header_cell.font = header_font
        header_cell.alignment = header_alignment
        header_cell.border = header_border

        values = [str(header_cell.value or "")]
        for row_index in range(2, min(worksheet.max_row, 200) + 1):
            values.append(str(worksheet.cell(row=row_index, column=column_index).value or ""))
        width = min(max(max((len(value) for value in values), default=0) + 2, 10), 48)
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    worksheet.row_dimensions[1].height = 30


def dataframe_to_excel_bytes(rows: list[dict[str, Any]], columns: list[str] | None = None) -> bytes:
    buffer = io.BytesIO()
    df = pd.DataFrame(rows)
    if columns:
        df = df.reindex(columns=columns)
    if "S No." in df.columns:
        df["S No."] = range(1, len(df) + 1)

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data")
        style_output_worksheet(writer.sheets["Data"])
    buffer.seek(0)
    return buffer.getvalue()


def create_district_zip(district: str, rows: list[dict[str, Any]]) -> bytes:
    zip_buffer = io.BytesIO()
    district_stem = safe_stem(district)
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as district_zip:
        district_zip.writestr(f"{district_stem}.xlsx", dataframe_to_excel_bytes(rows))
    zip_buffer.seek(0)
    return zip_buffer.getvalue()


def create_master_zip(
    district_data: dict[str, list[dict[str, Any]]],
    errors: list[dict[str, Any]],
    summary: dict[str, Any],
) -> bytes:
    buffer = io.BytesIO()
    columns = summary.get("columns") if isinstance(summary.get("columns"), list) else None
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as master_zip:
        used_names: set[str] = set()
        for district, rows in ordered_group_items(district_data):
            stem = safe_stem(district)
            filename = f"{stem}.xlsx"
            index = 2
            while filename in used_names:
                filename = f"{stem} {index}.xlsx"
                index += 1
            used_names.add(filename)
            master_zip.writestr(filename, dataframe_to_excel_bytes(rows, columns))
    buffer.seek(0)
    return buffer.getvalue()


