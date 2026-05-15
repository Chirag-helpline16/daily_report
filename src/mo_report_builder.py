import argparse
import csv
import os
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


SOURCE_FILE = Path("AdditionalInformationReport _ 12_05_2026 12_32_02.csv")
THREAD_ID = os.environ.get("CODEX_THREAD_ID", "mo-report")
OUTPUT_DIR = Path("outputs") / THREAD_ID
DEFAULT_ACK_PREFIX = "311"
DEFAULT_OUTPUT_NAME = f"MO_Unique_Report_{DEFAULT_ACK_PREFIX}.xlsx"
OUTPUT_FILE = OUTPUT_DIR / DEFAULT_OUTPUT_NAME
SUPPORTED_SOURCE_EXTENSIONS = {".csv", ".xlsx", ".xlsm"}


def clean_mo(value):
    text = (value or "").upper()
    text = text.replace("\u20b9", " ")

    # Remove operational annotations and amount fragments.
    replacements = [
        (r"GOLD[AE]N\s*(?:HOURS?|HR)\s*HEL[P]?LINE\s*\d*", " "),
        (r"GOLD[AE]N\s*(?:HOURS?|HR)", " "),
        (r"LIVE\s*HEL[P]?LINE\s*\d*", " "),
        (r"LIVEHEL[P]?LINE\s*\d*", " "),
        (r"(?:A\s*M|P\s*M|AM|PM)\s*HEL[P]?LINE\s*\d*", " "),
        (r"HEL[P]?LINE\s*\d*|HELPINE\s*\d*", " "),
        (r"\(\s*LIVE[^)]*\)", " "),
        (r"\bLIVE\b\s*[-:]?\s*\d{0,2}:?\d{0,2}\s*(?:AM|PM)?", " "),
        (r"\b\d{1,2}:\d{2}\s*(?:AM|PM)?\b", " "),
        (r"\(\s*\d{1,2}:\d{2}\s*(?:AM|PM)?\s*\)", " "),
        (
            r"\b(?:WHATS\s*APP|WHATSAPP|WHATSAAP|WHTSAPP|WAPP|WP|W|WH|WHAT)\s*(?:NO|NUMBER)\b",
            " ",
        ),
        (r"\b(?:MOBILE|PHONE|CONTACT)\s*(?:NO|NUMBER)\b", " "),
        (r"\b(?:AMOUNT|AMT)\s*[-:]?", " "),
        (r"RS\.?|INR|RUPEES?|/-|USD|USDT|TRT", " "),
        (r"\b\d[\d,\s]*(?:\.\d+)?\b", " "),
        (r"[^A-Z]+", " "),
        (r"\bP\s*M\b|\bA\s*M\b|\bPM\b|\bAM\b", " "),
    ]
    for pattern, repl in replacements:
        text = re.sub(pattern, repl, text)

    corrections = {
        "FROUD": "FRAUD",
        "FRAD": "FRAUD",
        "FRUAD": "FRAUD",
        "FRAUDS": "FRAUD",
        "FRAUDE": "FRAUD",
        "FRAUDD": "FRAUD",
        "FROAD": "FRAUD",
        "FRAYD": "FRAUD",
        "FACK": "FAKE",
        "FAK": "FAKE",
        "FAKIE": "FAKE",
        "FSKE": "FAKE",
        "FAHE": "FAKE",
        "EIDENTITY": "IDENTITY",
        "IDENTY": "IDENTITY",
        "IDENTITYY": "IDENTITY",
        "IDNTITY": "IDENTITY",
        "IDENTIY": "IDENTITY",
        "IDETITY": "IDENTITY",
        "IDIENTITY": "IDENTITY",
        "IDENITITY": "IDENTITY",
        "IDENITY": "IDENTITY",
        "IDENETITY": "IDENTITY",
        "IDENITTY": "IDENTITY",
        "IEDNTITY": "IDENTITY",
        "IDENTITYA": "IDENTITY",
        "IDENTIYA": "IDENTITY",
        "IDEENTITY": "IDENTITY",
        "IDEMNTITY": "IDENTITY",
        "IDENTITIY": "IDENTITY",
        "IDNETITY": "IDENTITY",
        "IDENTITUY": "IDENTITY",
        "INDENTITY": "IDENTITY",
        "IDFENTITY": "IDENTITY",
        "IIDENTITY": "IDENTITY",
        "IDENTTIY": "IDENTITY",
        "IDEINTITY": "IDENTITY",
        "FAEK": "FAKE",
        "INVESTMANT": "INVESTMENT",
        "INVESMENT": "INVESTMENT",
        "INVESTEMENT": "INVESTMENT",
        "INVESTEMNT": "INVESTMENT",
        "INVESTMET": "INVESTMENT",
        "INVESTMETN": "INVESTMENT",
        "INVETMENT": "INVESTMENT",
        "INVESRMENT": "INVESTMENT",
        "INVESTMEN": "INVESTMENT",
        "INVESTMRNT": "INVESTMENT",
        "INVETSMENT": "INVESTMENT",
        "INVESTMENTT": "INVESTMENT",
        "INVTMENT": "INVESTMENT",
        "HARASMENT": "HARASSMENT",
        "HARRASMENT": "HARASSMENT",
        "HARRESMENT": "HARASSMENT",
        "HARESSMENT": "HARASSMENT",
        "HARASSMNET": "HARASSMENT",
        "HARRASING": "HARASSING",
        "TRADNING": "TRADING",
        "TRADDING": "TRADING",
        "TRAIDING": "TRADING",
        "TRAVELLING": "TRAVELING",
        "WATHOUT": "WITHOUT",
        "WITHOT": "WITHOUT",
        "WITHOTU": "WITHOUT",
        "WITOUT": "WITHOUT",
        "WIHTOUT": "WITHOUT",
        "WITJH": "WITH",
        "WIYH": "WITH",
        "WITHAOUT": "WITHOUT",
        "WIHOUT": "WITHOUT",
        "WTHOUT": "WITHOUT",
        "WHITOUT": "WITHOUT",
        "WITHPUT": "WITHOUT",
        "WITHUT": "WITHOUT",
        "WITHOU": "WITHOUT",
        "WOTHOUT": "WITHOUT",
        "WTIHOUT": "WITHOUT",
        "WITHOIUT": "WITHOUT",
        "WITHGOUT": "WITHOUT",
        "WITHIUT": "WITHOUT",
        "WITHOUR": "WITHOUT",
        "WITHOUTR": "WITHOUT",
        "WITHOUTB": "WITHOUT",
        "WIRHOUT": "WITHOUT",
        "WITHP": "WITHOUT",
        "WITHTOUT": "WITHOUT",
        "WITHOUTP": "WITHOUT",
        "OPTP": "OTP",
        "OPT": "OTP",
        "OTO": "OTP",
        "OTRP": "OTP",
        "TOTP": "OTP",
        "SHOPING": "SHOPPING",
        "SHOPPINGG": "SHOPPING",
        "SHPPING": "SHOPPING",
        "SHOOPING": "SHOPPING",
        "SHOOPPING": "SHOPPING",
        "SHOPPINH": "SHOPPING",
        "SHOPPIG": "SHOPPING",
        "SHOPPIING": "SHOPPING",
        "SHOPPONG": "SHOPPING",
        "SHOPPNG": "SHOPPING",
        "SHOPPIN": "SHOPPING",
        "SHOPPIG": "SHOPPING",
        "BOOKINGS": "BOOKING",
        "BOOKIG": "BOOKING",
        "BOOKIMG": "BOOKING",
        "BOKING": "BOOKING",
        "BOOOKING": "BOOKING",
        "TIKIT": "TICKET",
        "TICKETS": "TICKET",
        "AP": "APP",
        "APPS": "APP",
        "APPLICATIONS": "APPLICATION",
        "WHTSAPP": "WHATSAPP",
        "WHATSAP": "WHATSAPP",
        "WHATSAAP": "WHATSAPP",
        "ADHAR": "AADHAR",
        "AADHHAR": "AADHAR",
        "AADHAAR": "AADHAR",
        "LONE": "LOAN",
        "REPIDO": "RAPIDO",
        "ESCOURT": "ESCORT",
        "ESCURT": "ESCORT",
        "ESCOERT": "ESCORT",
        "BLAKMAILING": "BLACKMAILING",
        "BLAKMAILNG": "BLACKMAILING",
        "BLACKMALING": "BLACKMAILING",
        "BLECKMAIL": "BLACKMAIL",
        "NAIGERIAN": "NIGERIAN",
        "NAIGERIYAN": "NIGERIAN",
        "NAIGERAIN": "NIGERIAN",
        "NAIZERIAN": "NIGERIAN",
        "NIGERIYAN": "NIGERIAN",
        "LOTTRY": "LOTTERY",
        "SCM": "SCAM",
        "DIGTAL": "DIGITAL",
        "GOVT": "GOV",
        "CHALAN": "CHALLAN",
        "CRAD": "CARD",
        "CREIDT": "CREDIT",
        "CREDI": "CREDIT",
        "DEDIT": "DEBIT",
        "DARD": "CARD",
        "CAR": "CARD",
        "LIMK": "LINK",
        "LIK": "LINK",
        "LNK": "LINK",
        "BI": "BY",
        "BYU": "BY",
        "VIDIO": "VIDEO",
        "PIC": "PHOTO",
        "CION": "COIN",
        "GIFTCARD": "GIFT CARD",
        "CASH": "CASH",
        "MATROMONIAL": "MATRIMONIAL",
        "MATRIMONY": "MATRIMONIAL",
        "AUSTROLOGY": "ASTROLOGY",
        "ASTROLOGER": "ASTROLOGY",
        "TACK": "TASK",
        "TAKS": "TASK",
        "TASKFRAUD": "TASK FRAUD",
        "SWAPING": "SWAP",
        "DATTING": "DATING",
        "FOLLOWRES": "FOLLOWERS",
        "FOLLOWE": "FOLLOWERS",
        "ELEVAN": "ELEVEN",
        "FEES": "FEE",
        "INSTA": "INSTAGRAM",
        "HACKE": "HACK",
    }
    words = []
    for word in text.split():
        replacement = corrections.get(word, word)
        words.extend(replacement.split())

    noise = {"SM", "WN", "WP", "WAPP", "VCT", "NO", "NUMBER", "H", "HR", "CALL", "SMS"}
    words = [word for word in words if word not in noise]
    text = " ".join(words)

    text = re.sub(r"\bWITH\s+OUT\b", "WITHOUT", text)
    text = re.sub(r"\bW\s+ITHOUT\b", "WITHOUT", text)
    text = re.sub(r"\bWIT\s+HOUT\b", "WITHOUT", text)
    text = re.sub(r"\bWITHOUT\s+OUT\s+OTP\b", "WITHOUT OTP", text)
    text = re.sub(r"\bWITHOUTOTP\b", "WITHOUT OTP", text)
    text = re.sub(r"\bWITHOTP\b", "WITH OTP", text)
    text = re.sub(r"\bIDENTITYFRAUD\b", "IDENTITY FRAUD", text)
    text = re.sub(r"\bTASKFRAUD\b", "TASK FRAUD", text)
    text = re.sub(r"\bCASH\s+BACK\b", "CASHBACK", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def contains(text, *needles):
    return any(needle in text for needle in needles)


def classify_mo(cleaned):
    text = f" {cleaned} "

    if contains(text, " RTO ", " CHALLAN ", " E CHALLAN "):
        return "RTO E-Challan Fraud"
    if contains(text, " DIGITAL ARREST "):
        return "Digital Arrest Fraud"
    if contains(text, " WITH OTP ") and " WITHOUT OTP " not in text:
        return "With OTP Fraud"
    if contains(text, " WITH FRAUD ") and " WITHOUT " not in text:
        return "With OTP Fraud"
    if contains(text, " WITHOUT OTP ", " WITHOUT FRAUD "):
        return "Without OTP Fraud"
    if (" FAKE " in text and "IDENT" in text) or contains(text, " FAKE ID ", " ID FAKE "):
        return "Fake Identity Fraud"
    if contains(text, " FAKE WEBSITE ", " WEBSITE FRAUD "):
        return "Fake Website Fraud"
    if contains(text, " AEPS ", " AADHAR "):
        return "Aadhaar AEPS Fraud"
    if contains(text, " CREDIT CARD ", " DEBIT CARD "):
        return "Credit/Debit Card Fraud"
    if contains(text, " CUSTOMER CARE ", " CUSTMER CARE ", " CARE FRAUD "):
        return "Customer Care Fraud"
    if contains(text, " KYC "):
        return "KYC Fraud"
    if contains(text, " GOV ", " GOVERNMENT ", " PMMVY "):
        return "Government Scheme Scam/Fraud"
    if contains(text, " BLACK MAIL", " BLACKMAIL", " BLACKMAILING ", " NUDE ", " VIDEO VIRAL ", " PHOTO VIRAL ", " SEXTORTION "):
        return "Blackmailing/Sextortion Fraud"
    if contains(text, " ESCORT "):
        return "Escort Service Fraud"
    if contains(text, " DATING APP ", " DATING FRAUD "):
        return "Dating App Fraud"
    if contains(text, " LOAN ") and contains(text, " HARASSMENT ", " HARASSING "):
        return "Loan Harassment/Fraud"
    if contains(text, " LOAN ", " LOAN APPLICATION "):
        return "Loan Fraud"
    if contains(text, " INVESTMENT ", " TRADING ", " NOVABIT ", " NOVA BIT ", " CRYPTO ", " COIN DCX ", " STOCK "):
        return "Investment Fraud"
    if contains(text, " SHOPPING ", " ORDER ", " REFUND "):
        return "Shopping Fraud"
    if contains(text, " JOB ", " PENCIL "):
        return "Job Fraud"
    if cleaned == "JO FRAUD":
        return "Job Fraud"
    if contains(text, " RAPIDO "):
        return "Rapido Fraud"
    if contains(text, " OLX "):
        return "OLX Fraud"
    if contains(text, " BOOKING ", " TICKET ", " HOTEL ", " RESORT ", " TRAVELING ", " TRAVELLING ", " TRAVEL ", " TRAVELS ", " TOUR ", " DWARKA ", " KEDARNATH ", " IPL "):
        return "Booking/Ticket Fraud"
    if contains(text, " TASK "):
        return "Task Fraud"
    if contains(text, " OLD COIN ", " OLD COINE ", " OLD CURRENCY "):
        return "Old Coin/Currency Fraud"
    if contains(text, " FRAUD BY LINK ", " LINK BY FRAUD ", " BY LINK ", " FRAUD BY FRAUD ", " APK FILE ", " APK "):
        return "Fraud by Link/APK"
    if contains(text, " WHATSAPP ") and contains(text, " HACK ", " HACKING "):
        return "WhatsApp Hack"
    if contains(text, " MOBILE THEFT ", " PHONE THEFT "):
        return "Mobile Theft Fraud"
    if contains(text, " INSTAGRAM ID HACK ", " INSTAGRAM HACK ", " PHONE HACK ", " MOBILE HACK ", " PHONE HACKING ", " MOBILE HACKING ", " HACK ", " HACKING "):
        return "Phone/Social Media Hack"
    if contains(text, " LOTTERY "):
        return "Lottery Fraud"
    if contains(text, " GIFT CARD ", " GIFT ", " REWARD ", " VOUCHER "):
        return "Gift/Reward Fraud"
    if contains(text, " CASHBACK "):
        return "Cashback Fraud"
    if contains(text, " INSURANCE ", " POLICY "):
        return "Insurance Policy Fraud"
    if contains(text, " VISA "):
        return "Visa Fraud"
    if contains(text, " MATRIMONIAL "):
        return "Matrimonial Fraud"
    if contains(text, " NIGERIAN "):
        return "Nigerian Fraud"
    if contains(text, " ASTROLOGY "):
        return "Astrology Fraud"
    if contains(text, " UPI "):
        return "UPI Fraud"
    if contains(text, " SIM SWAP "):
        return "SIM Swap Fraud"
    if contains(text, " ATM "):
        return "ATM Fraud"
    if contains(text, " NET BANKING ", " BOB BANK ", " BANK APP ", " IMPS "):
        return "Bank App/Net Banking Fraud"
    if contains(text, " FRAUD BY APP ", " BY APP ", " FAKE APPLICATION ", " APPLICATION FRAUD "):
        return "Fraud By App"
    if contains(text, " BILL PAYMENT ", " GAS BILL ", " GAS CYLINDER "):
        return "Utility/Bill Payment Fraud"
    if contains(text, " SOFTWARE UPDATE "):
        return "Software Update Fraud"
    if contains(text, " TRANSPORT "):
        return "Transport Fraud"
    if contains(text, " INSTAGRAM FAKE FOLLOWERS ", " FOLLOWERS FRAUD ", " INSTAGRAM FOLLOWERS "):
        return "Social Media Followers Fraud"
    if contains(text, " DREAM ELEVEN ", " DREAM FRAUD ", " FANTASY ", " GAMING "):
        return "Fantasy Sports/Gaming Fraud"
    if contains(text, " FRANCHISE "):
        return "Franchise Fraud"
    if contains(text, " BAJAJ FINANCE "):
        return "Finance Company Fraud"
    if contains(text, " HARASSMENT ", " OTP HARASSMENT ", " DIGITAL HARASSMENT "):
        return "Digital/OTP Harassment"
    if cleaned in {"OTHER", "OTHER FRAUD", "UNCLEAR OTHER", "UNCLEAR", ""}:
        return "Unclear/Other"
    return cleaned.title()


def as_number(value):
    if value is None:
        return None
    text = str(value).replace(",", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def autosize(sheet, widths):
    for col_idx, width in widths.items():
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def add_table(sheet, name, ref):
    table = Table(displayName=name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    sheet.add_table(table)


def header_key(value):
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def row_get(row, *names, default=""):
    for name in names:
        if name in row:
            return row.get(name, default)

    normalized = {header_key(key): value for key, value in row.items()}
    for name in names:
        key = header_key(name)
        if key in normalized:
            return normalized[key]
    return default


def validate_source_headers(headers):
    keys = {header_key(header) for header in headers if header}
    required_groups = {
        "Acknowledgement No.": [
            "Acknowledgement No.",
            "Acknowledgement No",
            "Acknowledgment No.",
            "Acknowledgment No",
        ],
        "Crime Aditional Information": [
            "Crime Aditional Information",
            "Crime Additional Information",
            "Additional Information",
        ],
    }

    missing = [
        label
        for label, aliases in required_groups.items()
        if not any(header_key(alias) in keys for alias in aliases)
    ]
    if missing:
        raise ValueError(f"Missing required column(s): {', '.join(missing)}")


def iter_csv_rows(source_file):
    for encoding in ("utf-8-sig", "utf-16", "cp1252"):
        try:
            with source_file.open(newline="", encoding=encoding) as handle:
                reader = csv.DictReader(handle)
                if not reader.fieldnames:
                    raise ValueError("The file does not contain a header row.")
                validate_source_headers(reader.fieldnames)
                rows = list(reader)
                yield from rows
            return
        except UnicodeError:
            continue

    raise ValueError("Could not read the CSV file. Please save it as UTF-8 CSV or Excel .xlsx.")


def iter_excel_rows(source_file):
    workbook = load_workbook(source_file, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        row_iter = sheet.iter_rows(values_only=True)
        headers = next(row_iter, None)
        if not headers:
            raise ValueError("The workbook does not contain a header row.")

        headers = [cell_text(header) for header in headers]
        validate_source_headers(headers)

        for values in row_iter:
            if not values or all(cell_text(value) == "" for value in values):
                continue
            yield {header: value for header, value in zip(headers, values) if header}
    finally:
        workbook.close()


def iter_source_rows(source_file):
    source_file = Path(source_file)
    suffix = source_file.suffix.lower()
    if suffix == ".csv":
        yield from iter_csv_rows(source_file)
        return
    if suffix in {".xlsx", ".xlsm"}:
        yield from iter_excel_rows(source_file)
        return
    raise ValueError("Unsupported file type. Upload a CSV, XLSX, or XLSM file.")


def build_report(
    source_file=SOURCE_FILE,
    output_file=None,
    output_dir=None,
    acknowledgement_prefix=DEFAULT_ACK_PREFIX,
):
    source_file = Path(source_file)
    if not source_file.exists():
        raise FileNotFoundError(f"Source file not found: {source_file}")

    if source_file.suffix.lower() not in SUPPORTED_SOURCE_EXTENSIONS:
        raise ValueError("Unsupported file type. Upload a CSV, XLSX, or XLSM file.")

    acknowledgement_prefix = cell_text(acknowledgement_prefix)
    safe_prefix = re.sub(r"[^A-Za-z0-9_-]+", "", acknowledgement_prefix) or "all"
    if output_file is None:
        output_dir = Path(output_dir) if output_dir else OUTPUT_DIR
        output_file = output_dir / f"MO_Unique_Report_{safe_prefix}.xlsx"
    else:
        output_file = Path(output_file)

    include_label = (
        f"Acknowledgement No. starts with {acknowledgement_prefix}"
        if acknowledgement_prefix
        else "All rows"
    )
    percent_header = f"% of {acknowledgement_prefix} Rows" if acknowledgement_prefix else "% of Rows"

    rows = []
    excluded_count = 0
    total_rows = 0

    for source_row in iter_source_rows(source_file):
        total_rows += 1
        ack = cell_text(
            row_get(
                source_row,
                "Acknowledgement No.",
                "Acknowledgement No",
                "Acknowledgment No.",
                "Acknowledgment No",
            )
        )
        if acknowledgement_prefix and not ack.startswith(acknowledgement_prefix):
            excluded_count += 1
            continue

        raw_info = cell_text(
            row_get(
                source_row,
                "Crime Aditional Information",
                "Crime Additional Information",
                "Additional Information",
            )
        )
        cleaned = clean_mo(raw_info)
        canonical = classify_mo(cleaned)
        rows.append(
            {
                "Acknowledgement No.": ack,
                "Crime Aditional Information": raw_info,
                "Cleaned MO Text": cleaned,
                "Final MO": canonical,
                "Category": cell_text(row_get(source_row, "Category")),
                "Sub Category": cell_text(row_get(source_row, "Sub Category", "Subcategory")),
                "Status": cell_text(row_get(source_row, "Status")),
                "Amount": as_number(row_get(source_row, "Amount")),
            }
        )

    mo_counts = Counter(row["Final MO"] for row in rows)
    variant_counts = defaultdict(Counter)
    examples = {}
    amount_totals = defaultdict(float)
    amount_counts = defaultdict(int)

    for row in rows:
        mo = row["Final MO"]
        variant_counts[mo][row["Cleaned MO Text"]] += 1
        examples.setdefault(mo, row["Crime Aditional Information"])
        if row["Amount"] is not None:
            amount_totals[mo] += row["Amount"]
            amount_counts[mo] += 1

    summary_rows = []
    for rank, (mo, count) in enumerate(mo_counts.most_common(), start=1):
        variants = variant_counts[mo].most_common(8)
        variant_text = "; ".join(f"{label or '[blank]'} ({variant_count})" for label, variant_count in variants)
        if len(variant_counts[mo]) > 8:
            variant_text += f"; +{len(variant_counts[mo]) - 8} more"
        avg_amount = amount_totals[mo] / amount_counts[mo] if amount_counts[mo] else None
        summary_rows.append(
            [
                rank,
                mo,
                count,
                count / len(rows) if rows else 0,
                len(variant_counts[mo]),
                round(amount_totals[mo], 2) if amount_counts[mo] else None,
                round(avg_amount, 2) if avg_amount is not None else None,
                variant_text,
                examples.get(mo, ""),
            ]
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "MO Summary"
    detail = wb.create_sheet("Detail Mapping")
    rules = wb.create_sheet("Rules")

    header_fill = PatternFill("solid", fgColor="1F4E78")
    title_fill = PatternFill("solid", fgColor="EAF2F8")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    thin = Side(style="thin", color="D9E2EC")
    border = Border(bottom=thin)

    ws["A1"] = f"MO Unique Report - {include_label}"
    ws["A1"].font = Font(bold=True, size=16, color="1F4E78")
    ws.merge_cells("A1:I1")
    ws["A2"] = f"Source: {source_file.name}"
    ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A5"] = "Metric"
    ws["B5"] = "Value"
    ws["A5"].fill = header_fill
    ws["B5"].fill = header_fill
    ws["A5"].font = white_font
    ws["B5"].font = white_font

    metrics = [
        ("Total rows in source", total_rows),
        (f"Rows counted ({include_label})", len(rows)),
        ("Rows ignored (all other acknowledgement numbers)", excluded_count),
        ("Unique MO categories identified", len(mo_counts)),
    ]
    for idx, metric in enumerate(metrics, start=6):
        ws.cell(idx, 1).value = metric[0]
        ws.cell(idx, 2).value = metric[1]
        ws.cell(idx, 1).font = bold_font
        ws.cell(idx, 1).fill = title_fill
        ws.cell(idx, 2).fill = title_fill

    start_row = 12
    headers = [
        "Rank",
        "Final MO",
        "Count",
        percent_header,
        "Grouped Variant Count",
        "Total Amount",
        "Average Amount",
        "Grouped Cleaned Variants",
        "Example Original Text",
    ]
    ws.append([])
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, data in enumerate(summary_rows, start=start_row + 1):
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row_idx, col_idx)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=col_idx in {8, 9})

    if summary_rows:
        last_row = start_row + len(summary_rows)
        add_table(ws, "MOSummaryTable", f"A{start_row}:I{last_row}")
        ws.freeze_panes = "A13"
        ws.auto_filter.ref = f"A{start_row}:I{last_row}"
        for cell in ws[f"D{start_row + 1}:D{last_row}"]:
            cell[0].number_format = "0.0%"
        for col in ("F", "G"):
            for cell in ws[f"{col}{start_row + 1}:{col}{last_row}"]:
                cell[0].number_format = "#,##0"

        chart_rows = min(15, len(summary_rows))
        chart = BarChart()
        chart.title = "Top MO Counts"
        chart.y_axis.title = "Cases"
        chart.x_axis.title = "MO"
        chart.height = 8
        chart.width = 16
        data = Reference(ws, min_col=3, min_row=start_row, max_row=start_row + chart_rows)
        cats = Reference(ws, min_col=2, min_row=start_row + 1, max_row=start_row + chart_rows)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.legend = None
        ws.add_chart(chart, "K5")

    autosize(ws, {1: 9, 2: 31, 3: 12, 4: 14, 5: 18, 6: 15, 7: 15, 8: 60, 9: 72, 11: 2})
    ws.row_dimensions[1].height = 24
    for row_idx in range(start_row + 1, start_row + len(summary_rows) + 1):
        ws.row_dimensions[row_idx].height = 45

    detail_headers = [
        "Acknowledgement No.",
        "Final MO",
        "Cleaned MO Text",
        "Crime Aditional Information",
        "Category",
        "Sub Category",
        "Status",
        "Amount",
    ]
    detail.append(detail_headers)
    for cell in detail[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in rows:
        detail.append([row[header] for header in detail_headers])
    if rows:
        add_table(detail, "DetailMappingTable", f"A1:H{len(rows) + 1}")
        detail.freeze_panes = "A2"
        detail.auto_filter.ref = f"A1:H{len(rows) + 1}"
        for cell in detail[f"H2:H{len(rows) + 1}"]:
            cell[0].number_format = "#,##0"
    autosize(detail, {1: 19, 2: 28, 3: 36, 4: 90, 5: 24, 6: 32, 7: 16, 8: 13})
    for row_cells in detail.iter_rows(min_row=2, max_row=min(len(rows) + 1, 600), min_col=1, max_col=8):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column in {3, 4})

    rules["A1"] = "Rules Used"
    rules["A1"].font = Font(bold=True, size=16, color="1F4E78")
    rules["A3"] = "Included rows"
    rules["B3"] = (
        f"Only records where Acknowledgement No. starts with {acknowledgement_prefix}"
        if acknowledgement_prefix
        else "All records"
    )
    rules["A4"] = "Ignored rows"
    rules["B4"] = (
        "All other acknowledgement numbers"
        if acknowledgement_prefix
        else "None filtered by acknowledgement number"
    )
    rules["A5"] = "Noise removed"
    rules["B5"] = "LIVE/time text, HELPLINE, GOLDEN HOUR(S), amounts, standalone AM/PM, phone/contact markers"
    rules["A6"] = "Spelling normalization"
    rules["B6"] = "Common human-entry variants were grouped, e.g. WITHOTU/WIT HOUT -> Without, FRAYD -> Fraud, NAIGERAIN/NAIZERIAN -> Nigerian, ADHAR/AADHAAR -> Aadhaar"
    rules["A7"] = "Audit trail"
    rules["B7"] = "Detail Mapping sheet shows the original text, cleaned MO text, and final grouped MO for each counted 311 row"
    rules["A9"] = "Important note"
    rules["B9"] = "This is a rule-assisted human-style grouping of short MO labels. Rows with unclear text are kept as Unclear/Other instead of being forced into a category."

    for row in range(3, 10):
        rules.cell(row, 1).font = bold_font
        rules.cell(row, 1).fill = title_fill
        rules.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
    autosize(rules, {1: 24, 2: 115})

    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)

    # Compact verification pass.
    check_wb = load_workbook(output_file, data_only=True)
    check_summary = check_wb["MO Summary"]
    counted = check_summary["B7"].value
    unique = check_summary["B9"].value
    check_wb.close()
    if counted != len(rows):
        raise RuntimeError(f"Count verification failed: workbook={counted}, expected={len(rows)}")
    if unique != len(mo_counts):
        raise RuntimeError(f"Unique MO verification failed: workbook={unique}, expected={len(mo_counts)}")

    return {
        "output_file": output_file.resolve(),
        "source_file": source_file.name,
        "total_rows": total_rows,
        "rows_counted": len(rows),
        "rows_ignored": excluded_count,
        "unique_mo_categories": len(mo_counts),
        "top_10": [{"mo": mo, "count": count} for mo, count in mo_counts.most_common(10)],
    }


def main():
    parser = argparse.ArgumentParser(description="Build an MO unique report from a CSV/XLSX source file.")
    parser.add_argument("source", nargs="?", default=SOURCE_FILE, help="Source CSV, XLSX, or XLSM file")
    parser.add_argument("--prefix", default=DEFAULT_ACK_PREFIX, help="Acknowledgement number prefix to include")
    parser.add_argument("--output", default=None, help="Output .xlsx file path")
    args = parser.parse_args()

    result = build_report(args.source, output_file=args.output, acknowledgement_prefix=args.prefix)

    print(f"Saved: {result['output_file']}")
    print(f"Rows counted: {result['rows_counted']}")
    print(f"Unique MO categories: {result['unique_mo_categories']}")
    print("Top 10:")
    for item in result["top_10"]:
        print(f"{item['count']}\t{item['mo']}")


if __name__ == "__main__":
    main()
