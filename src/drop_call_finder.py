"""Drop call finder page and workbook processor.

Ports the DROP CALLS Excel processor into the Streamlit app without writing
uploaded or generated files to project folders.
"""

from __future__ import annotations

import io
from dataclasses import dataclass

import pandas as pd
import streamlit as st
from openpyxl.styles import Font


REQUIRED_COLUMNS = {"phone_number_dialed", "status"}


@dataclass(frozen=True)
class DropCallResult:
    """Processed workbook payload and summary."""

    workbook: bytes
    filename: str
    original_rows: int
    unique_phone_numbers: int
    drop_only_numbers: int
    status_columns: list[str]


def generate_filename_from_dates(df: pd.DataFrame) -> str:
    """Generate filename based on earliest and latest dates from call_date column."""
    try:
        if "call_date" not in df.columns:
            return "processed_output.xlsx"

        if pd.api.types.is_datetime64_any_dtype(df["call_date"]):
            df["call_date_parsed"] = pd.to_datetime(df["call_date"], errors="coerce")
        else:
            df["call_date_parsed"] = pd.to_datetime(
                df["call_date"],
                format="%d-%m-%Y %H:%M:%S",
                errors="coerce",
            )
            if df["call_date_parsed"].isna().all() and not df["call_date"].isna().all():
                df["call_date_parsed"] = pd.to_datetime(
                    df["call_date"],
                    dayfirst=True,
                    errors="coerce",
                )

        valid_dates = df["call_date_parsed"].dropna()
        if len(valid_dates) == 0:
            return "processed_output.xlsx"

        earliest_date = valid_dates.min()
        latest_date = valid_dates.max()
        return f"{earliest_date.strftime('%d-%m-%Y')} to {latest_date.strftime('%d-%m-%Y')} call.xlsx"
    except Exception:
        return "processed_output.xlsx"


def _create_pivot(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Create the same pivot table used by the standalone tool."""
    pivot_with_counts = df.groupby(["phone_number_dialed", "status"]).size().unstack(fill_value=0)
    pivot = pivot_with_counts.copy().replace(0, pd.NA)
    pivot["Grand Total"] = pivot.notna().sum(axis=1)
    return pivot.reset_index(), pivot_with_counts


def create_pivot_sheet(df: pd.DataFrame, writer: pd.ExcelWriter) -> pd.DataFrame:
    """Create Sheet2 with pivot counts and a bold Grand Total row."""
    pivot, pivot_with_counts = _create_pivot(df)
    pivot.to_excel(writer, sheet_name="Sheet2", index=False)

    worksheet = writer.sheets["Sheet2"]
    status_cols = [col for col in pivot.columns if col != "phone_number_dialed"]
    last_row = len(pivot) + 2

    worksheet.cell(row=last_row, column=1, value="Grand Total")
    worksheet.cell(row=last_row, column=1).font = Font(bold=True)

    for idx, col in enumerate(status_cols, start=2):
        if col == "Grand Total":
            total = pivot_with_counts.sum().sum()
        else:
            total = pivot[col].sum()
        cell = worksheet.cell(row=last_row, column=idx, value=total)
        cell.font = Font(bold=True)

    return pivot


def create_filtered_sheet(df: pd.DataFrame, writer: pd.ExcelWriter) -> pd.DataFrame:
    """Create Sheet3 with phone numbers where DROP exists and all other statuses are blank."""
    pivot, _ = _create_pivot(df)

    status_columns = [
        col for col in pivot.columns
        if col not in ["phone_number_dialed", "Grand Total"]
    ]
    other_status_cols = [col for col in status_columns if col != "DROP"]

    mask = pivot["DROP"].notna()
    for col in other_status_cols:
        if col in pivot.columns:
            mask = mask & pivot[col].isna()

    df_filtered = pivot[mask].copy()
    df_filtered.to_excel(writer, sheet_name="Sheet3", index=False)
    return df_filtered


def create_phone_numbers_sheet(df_sheet3: pd.DataFrame, writer: pd.ExcelWriter) -> pd.DataFrame:
    """Create Sheet4 containing only serial number and phone_number_dialed."""
    df_phone = df_sheet3[["phone_number_dialed"]].copy()
    df_phone.insert(0, "sr", range(1, len(df_phone) + 1))
    df_phone.to_excel(writer, sheet_name="Sheet4", index=False)
    return df_phone


def process_drop_call_workbook(file_obj) -> DropCallResult:
    """Process uploaded Excel bytes and return an in-memory output workbook."""
    df_original = pd.read_excel(file_obj, sheet_name="Sheet1")
    missing = sorted(REQUIRED_COLUMNS - set(df_original.columns))
    if missing:
        missing_text = ", ".join(missing)
        raise ValueError(f"Sheet1 must contain required column(s): {missing_text}")

    download_filename = generate_filename_from_dates(df_original)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_original.to_excel(writer, sheet_name="Sheet1", index=False)
        pivot = create_pivot_sheet(df_original, writer)
        df_filtered = create_filtered_sheet(df_original, writer)
        create_phone_numbers_sheet(df_filtered, writer)

    return DropCallResult(
        workbook=output.getvalue(),
        filename=download_filename,
        original_rows=len(df_original),
        unique_phone_numbers=len(pivot),
        drop_only_numbers=len(df_filtered),
        status_columns=[
            str(col) for col in pivot.columns
            if col not in ["phone_number_dialed", "Grand Total"]
        ],
    )


def render_drop_call_finder_page():
    """Render the Drop Call Finder page."""
    st.title("Drop Call Finder")
    st.markdown(
        "Upload an Excel file with `Sheet1` data to generate the same 4-sheet drop-call workbook: "
        "original data, status pivot, DROP-only filtered rows, and phone numbers."
    )

    uploaded_file = st.file_uploader(
        "Choose an Excel file",
        type=["xlsx", "xls"],
        key="drop_call_finder_uploader",
        help="Sheet1 must contain phone_number_dialed and status columns.",
    )

    with st.expander("What this tool creates", expanded=False):
        st.markdown(
            """
            - **Sheet1:** Original uploaded data.
            - **Sheet2:** Pivot table with phone numbers as rows, status as columns, and Grand Total row.
            - **Sheet3:** Only phone numbers where `DROP` has a value and every other status is blank.
            - **Sheet4:** Serial number plus `phone_number_dialed` from Sheet3.
            """
        )

    if uploaded_file is None:
        st.info("Upload an Excel file to find DROP-only phone numbers.")
        return

    try:
        result = process_drop_call_workbook(uploaded_file)
    except Exception as exc:
        st.error(str(exc))
        return

    st.success("Drop call workbook generated successfully.")

    metric_cols = st.columns(4)
    metric_cols[0].metric("Original rows", f"{result.original_rows:,}")
    metric_cols[1].metric("Unique phones", f"{result.unique_phone_numbers:,}")
    metric_cols[2].metric("Drop-only phones", f"{result.drop_only_numbers:,}")
    metric_cols[3].metric("Statuses", f"{len(result.status_columns):,}")

    st.download_button(
        "Download Drop Call Workbook",
        data=result.workbook,
        file_name=result.filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
        key="drop_call_finder_download",
    )

    with st.expander("Detected status columns", expanded=False):
        if result.status_columns:
            st.write(", ".join(result.status_columns))
        else:
            st.info("No status columns found.")
