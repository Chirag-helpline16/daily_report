import io
import re
import zipfile
from datetime import datetime
from typing import Dict, List, Tuple

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


OUTPUT_COLUMNS = [
    "SR NO",
    "ACK NO",
    "IFSC CODE",
    "BANK NAME",
    "AC NO",
    "SUSPECT NAME",
    "SUSPECT ADDRESS",
    "SUSPECT MOBILE NUMBER",
    "SUSPECT LOCATION",
]

SOURCE_FIELDS = {
    "ACK NO": [
        "acknowledgementno",
        "acknowledgementnumber",
        "acknowledgmentno",
        "ackno",
        "acknumber",
        "ack",
    ],
    "IFSC CODE": [
        "ifsccode",
        "ifsc",
    ],
    "BANK NAME": [
        "bankfis",
        "bankfi",
        "bankname",
        "bank",
        "bankfinancialinstitution",
    ],
    "AC NO": [
        "accountno",
        "accountnumber",
        "bankaccountno",
        "bankaccountnumber",
        "acno",
        "accno",
        "account",
    ],
}

_ILLEGAL_EXCEL_CHARS_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")
_INVALID_FILENAME_CHARS_RE = re.compile(r'[<>:"/\\|?*\x00-\x1F]+')


def _normalize_header(value) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value).lower())


def _clean_text(value) -> str:
    if value is None:
        return ""
    try:
        if value != value:
            return ""
    except Exception:
        pass
    text = str(value).strip()
    if text.lower() in {"nan", "none", "<na>", "nat"}:
        return ""
    return _ILLEGAL_EXCEL_CHARS_RE.sub("", text)


def _clean_identifier(value) -> str:
    text = _clean_text(value)
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    if re.fullmatch(r"\d+(?:\.0+)?[eE]\+?\d+", text):
        try:
            return str(int(float(text)))
        except (OverflowError, ValueError):
            return text
    return text


def _account_key(value) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "", _clean_identifier(value)).upper()


def detect_columns(columns: List[str]) -> Dict[str, str]:
    normalized = {_normalize_header(column): column for column in columns}
    detected: Dict[str, str] = {}
    for field, aliases in SOURCE_FIELDS.items():
        for alias in aliases:
            if alias in normalized:
                detected[field] = normalized[alias]
                break
    return detected


def process_gujarat_account_file(
    df: pd.DataFrame,
    column_mapping: Dict[str, str],
) -> Tuple[pd.DataFrame, Dict[str, int]]:
    missing = [
        field
        for field in ("ACK NO", "IFSC CODE", "BANK NAME", "AC NO")
        if not column_mapping.get(field) or column_mapping[field] not in df.columns
    ]
    if missing:
        raise ValueError(f"Missing source column mapping for: {', '.join(missing)}")

    output = pd.DataFrame(
        {
            "ACK NO": df[column_mapping["ACK NO"]].map(_clean_identifier),
            "IFSC CODE": df[column_mapping["IFSC CODE"]].map(_clean_identifier),
            "BANK NAME": df[column_mapping["BANK NAME"]].map(_clean_text),
            "AC NO": df[column_mapping["AC NO"]].map(_clean_identifier),
            "SUSPECT NAME": "",
            "SUSPECT ADDRESS": "",
            "SUSPECT MOBILE NUMBER": "",
            "SUSPECT LOCATION": "",
        }
    )

    original_rows = len(output)
    output["_account_key"] = output["AC NO"].map(_account_key)
    blank_account_rows = int((output["_account_key"] == "").sum())
    output = output[output["_account_key"] != ""].copy()

    before_dedupe = len(output)
    output = output.drop_duplicates(subset="_account_key", keep="first").copy()
    duplicate_account_rows = before_dedupe - len(output)

    output = output.sort_values(
        by=["BANK NAME", "AC NO"],
        key=lambda series: series.astype(str).str.upper(),
        kind="stable",
    ).reset_index(drop=True)

    output.insert(0, "SR NO", range(1, len(output) + 1))
    output = output[OUTPUT_COLUMNS]

    stats = {
        "input_rows": original_rows,
        "blank_account_rows": blank_account_rows,
        "duplicate_account_rows": duplicate_account_rows,
        "output_rows": len(output),
        "bank_count": int(output["BANK NAME"].replace("", "UNKNOWN BANK").nunique()),
    }
    return output, stats


def _safe_sheet_name(value: str) -> str:
    name = _INVALID_FILENAME_CHARS_RE.sub("_", _clean_text(value)).strip(" .'_")
    return (name or "Data")[:31]


def _safe_filename(value: str) -> str:
    name = _INVALID_FILENAME_CHARS_RE.sub("_", _clean_text(value)).strip(" ._")
    return name or "UNKNOWN_BANK"


def _style_worksheet(worksheet, table_name: str) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Calibri", size=10, color="111827")
    odd_fill = PatternFill("solid", fgColor="FFFFFF")
    even_fill = PatternFill("solid", fgColor="F4F8FB")
    blank_fill = PatternFill("solid", fgColor="FFF7DA")
    border_side = Side(style="thin", color="D9E2F3")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    worksheet.row_dimensions[1].height = 28

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for row_index, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        row_fill = even_fill if row_index % 2 == 0 else odd_fill
        worksheet.row_dimensions[row_index].height = 24
        for column_index, cell in enumerate(row, start=1):
            cell.font = body_font
            cell.border = border
            cell.fill = blank_fill if column_index >= 6 else row_fill
            cell.alignment = center if column_index in {1, 2, 3, 5, 7} else left
            if column_index in {2, 3, 5, 7}:
                cell.number_format = "@"

    preferred_widths = {
        "A": 8,
        "B": 22,
        "C": 16,
        "D": 30,
        "E": 24,
        "F": 24,
        "G": 22,
        "H": 38,
        "I": 28,
    }
    for column_cells in worksheet.columns:
        letter = column_cells[0].column_letter
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[letter].width = min(max(preferred_widths.get(letter, 12), max_length + 3), 48)

    if worksheet.max_row > 1:
        table_ref = f"A1:I{worksheet.max_row}"
        worksheet.auto_filter.ref = table_ref
        table = Table(displayName=table_name, ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
    else:
        worksheet.auto_filter.ref = "A1:I1"


def dataframe_to_styled_excel_bytes(df: pd.DataFrame, sheet_name: str = "Unique Accounts") -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=_safe_sheet_name(sheet_name))
        worksheet = writer.sheets[_safe_sheet_name(sheet_name)]
        _style_worksheet(worksheet, "UniqueAccountsTable")
    return buffer.getvalue()


def bankwise_zip_bytes(df: pd.DataFrame) -> bytes:
    zip_buffer = io.BytesIO()
    used_names = set()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        grouped = df.assign(_bank=df["BANK NAME"].replace("", "UNKNOWN BANK")).groupby("_bank", sort=True)
        for bank_name, bank_df in grouped:
            bank_output = bank_df.drop(columns=["_bank"]).copy().reset_index(drop=True)
            bank_output["SR NO"] = range(1, len(bank_output) + 1)
            excel_bytes = dataframe_to_styled_excel_bytes(bank_output, sheet_name=str(bank_name))

            filename_root = _safe_filename(str(bank_name))
            filename = f"{filename_root}.xlsx"
            suffix = 2
            while filename.lower() in used_names:
                filename = f"{filename_root}_{suffix}.xlsx"
                suffix += 1
            used_names.add(filename.lower())
            archive.writestr(filename, excel_bytes)
    return zip_buffer.getvalue()


def _read_uploaded_excel(uploaded_file, sheet_name: str) -> pd.DataFrame:
    uploaded_file.seek(0)
    return pd.read_excel(uploaded_file, sheet_name=sheet_name, dtype=str, keep_default_na=False)


def _select_source_column(label: str, columns: List[str], detected: Dict[str, str]) -> str:
    options = ["-- Select --"] + columns
    detected_column = detected.get(label)
    index = options.index(detected_column) if detected_column in options else 0
    return st.selectbox(label, options=options, index=index, key=f"guj_formatter_{label}")


def render_gujarat_account_formatter_page():
    st.title("Gujarat Unique Account Output")

    uploaded_file = st.file_uploader(
        "Upload Gujarat Excel File",
        type=["xlsx", "xls"],
        key="gujarat_account_formatter_upload",
    )

    if uploaded_file is None:
        st.info("Upload the Gujarat Excel generated from Automated Workflow.")
        return

    try:
        file_bytes = uploaded_file.getvalue()
        excel_file = pd.ExcelFile(io.BytesIO(file_bytes))
        sheet_name = excel_file.sheet_names[0]
        if len(excel_file.sheet_names) > 1:
            sheet_name = st.selectbox(
                "Sheet",
                options=excel_file.sheet_names,
                key="gujarat_account_formatter_sheet",
            )

        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, dtype=str, keep_default_na=False)
        st.success(f"Loaded {len(df):,} rows from {uploaded_file.name}")

        detected = detect_columns(list(df.columns))
        st.markdown("### Source Columns")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            ack_col = _select_source_column("ACK NO", list(df.columns), detected)
        with col2:
            ifsc_col = _select_source_column("IFSC CODE", list(df.columns), detected)
        with col3:
            bank_col = _select_source_column("BANK NAME", list(df.columns), detected)
        with col4:
            account_col = _select_source_column("AC NO", list(df.columns), detected)

        selected = {
            "ACK NO": ack_col,
            "IFSC CODE": ifsc_col,
            "BANK NAME": bank_col,
            "AC NO": account_col,
        }
        if any(value == "-- Select --" for value in selected.values()):
            st.warning("Select all four source columns to generate output.")
            return

        output_df, stats = process_gujarat_account_file(df, selected)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_data = dataframe_to_styled_excel_bytes(output_df)
        zip_data = bankwise_zip_bytes(output_df) if not output_df.empty else b""

        st.markdown("### Output Summary")
        metric_cols = st.columns(5)
        metric_cols[0].metric("Input Rows", f"{stats['input_rows']:,}")
        metric_cols[1].metric("Output Rows", f"{stats['output_rows']:,}")
        metric_cols[2].metric("Duplicate AC Removed", f"{stats['duplicate_account_rows']:,}")
        metric_cols[3].metric("Blank AC Removed", f"{stats['blank_account_rows']:,}")
        metric_cols[4].metric("Banks", f"{stats['bank_count']:,}")

        st.dataframe(output_df.head(100), use_container_width=True, hide_index=True)

        dl_col1, dl_col2 = st.columns(2)
        with dl_col1:
            st.download_button(
                "Download Proper Excel",
                data=excel_data,
                file_name=f"Gujarat_Unique_Accounts_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
                key="gujarat_account_formatter_excel",
            )
        with dl_col2:
            st.download_button(
                "Download Bank-Wise ZIP",
                data=zip_data,
                file_name=f"Gujarat_Bank_Wise_Unique_Accounts_{timestamp}.zip",
                mime="application/zip",
                type="primary",
                use_container_width=True,
                disabled=output_df.empty,
                key="gujarat_account_formatter_zip",
            )
    except Exception as exc:
        st.error(f"Error processing file: {exc}")
